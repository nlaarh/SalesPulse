#!/usr/bin/env python3
"""
AAA WCNY — Membership Penetration Strategic Analysis
v4 + MAPS: County choropleth, city bubble map, old/new vehicle maps
Big Picture narrative, clean exec summary, Gartner-style design
"""
import openpyxl, io, os, base64, sqlite3, json
from datetime import datetime

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import matplotlib.colors as mcolors
from matplotlib.patches import Patch
from matplotlib.colors import to_rgba

# ── Palette (Deloitte/Accenture: clean, professional) ──────────────────
BG       = "#FFFFFF"
NAVY     = "#002B5C"
DARK     = "#1A3A5C"
TEAL     = "#00838F"
ACCENT   = "#86BC25"
GOLD     = "#D4A843"
LGRAY    = "#F5F6F8"
MGRAY    = "#B0BEC5"
DGRAY    = "#546E7A"
WHITE    = "#FFFFFF"
RED      = "#D32F2F"
GREEN    = "#388E3C"
ORANGE   = "#EF6C00"
BLUE     = "#1565C0"
PURPLE   = "#7B1FA2"
PINK     = "#C2185B"

PIE = ["#002B5C","#00838F","#86BC25","#D4A843","#EF6C00",
       "#7B1FA2","#1565C0","#C2185B","#388E3C","#D32F2F","#B0BEC5"]

ROOT = os.path.dirname(os.path.abspath(__file__))
TOTAL_MEMBERS = 843_000


def load():
    wb = openpyxl.load_workbook(os.path.join(ROOT,"AAA_WCNY_Territory_Data.xlsx"), read_only=True)
    n = lambda v: v if isinstance(v,(int,float)) else 0
    census={}
    for r in list(wb["Census Data by Zip"].iter_rows(values_only=True))[1:]:
        if r[0] and str(r[0])!="TOTAL":
            z=str(r[0]).zfill(5)
            census[z]=dict(pop=n(r[3]),adults=n(r[4]),income=n(r[5]),
                           age=n(r[6]),hh=n(r[7]),college=n(r[10]))
    dmv={}
    for r in list(wb["DMV Info"].iter_rows(values_only=True))[1:]:
        if r[0]:
            z=str(r[0]).zfill(5)
            dmv[z]=dict(vehicles=int(n(r[3])),old=int(n(r[4])),new=int(n(r[5])))
    mem={}
    for r in list(wb["Members & Customers by Zip"].iter_rows(values_only=True))[1:]:
        if r[0] and r[2] and n(r[3])>0:
            z=str(r[0]).zfill(5)
            mem[z]=dict(city=r[1] or "",county=r[2],members=n(r[3]),pop=n(r[9]))
    return mem,census,dmv


def load_geo():
    """Load county GeoJSON polygons and zip centroids from the salesinsight DB."""
    db = os.path.expanduser("~/.salesinsight/salesinsight.db")
    conn = sqlite3.connect(db)
    county_geo = {}
    for name, gj in conn.execute(
        "SELECT name, geojson FROM geo_counties WHERE geojson IS NOT NULL"
    ).fetchall():
        try:
            county_geo[name] = json.loads(gj)
        except Exception:
            pass
    zip_latlng = {}
    for row in conn.execute(
        "SELECT zip_code, city, county_name, lat, lng FROM geo_zips WHERE lat IS NOT NULL AND lng IS NOT NULL"
    ).fetchall():
        z, city, county, lat, lng = row
        if lat and lng:
            zip_latlng[str(z).zfill(5)] = {
                "city": city or "", "county": county or "",
                "lat": float(lat), "lng": float(lng)
            }
    conn.close()
    return county_geo, zip_latlng


def chart_county_choropleth(county_df, county_geo):
    """County-level penetration choropleth using GeoJSON polygon boundaries."""
    from matplotlib.patches import Polygon as MplPoly
    from matplotlib.collections import PatchCollection

    pen_map = dict(zip(county_df["county"], county_df["pen"]))
    cmap = mcolors.LinearSegmentedColormap.from_list("pen", [RED, ORANGE, GOLD, GREEN])
    vmax = max(pen_map.values()) if pen_map else 40

    fig, ax = plt.subplots(figsize=(14, 10), facecolor=WHITE)
    ax.set_facecolor("#D6EAF8")  # light-blue water background
    ax.set_aspect("equal")
    ax.axis("off")
    ax.set_title("AAA WCNY Territory — Member Penetration by County",
                 color=NAVY, fontsize=15, fontweight="bold", pad=16)

    plotted = []
    for name, geom in county_geo.items():
        pen = pen_map.get(name, 0)
        color = cmap(pen / vmax)
        coords = geom.get("coordinates", [])
        # Handle both Polygon and MultiPolygon
        rings = [coords[0]] if geom["type"] == "Polygon" else [p[0] for p in coords]
        for ring in rings:
            pts = [(x, y) for x, y in ring]
            if len(pts) < 3:
                continue
            poly = MplPoly(pts, closed=True)
            pc = PatchCollection([poly], facecolor=color,
                                 edgecolor=WHITE, linewidths=0.8, zorder=2)
            ax.add_collection(pc)
            plotted.append(pts)

        # Label county name + penetration at centroid
        if rings:
            xs = [p[0] for p in rings[0]]
            ys = [p[1] for p in rings[0]]
            cx, cy = sum(xs)/len(xs), sum(ys)/len(ys)
            ax.text(cx, cy, f"{name}\n{pen:.1f}%",
                    ha="center", va="center", fontsize=6.5,
                    color=WHITE if pen > vmax*0.5 else NAVY,
                    fontweight="bold", zorder=3)

    if plotted:
        all_x = [p[0] for pts in plotted for p in pts]
        all_y = [p[1] for pts in plotted for p in pts]
        pad_x = (max(all_x)-min(all_x))*0.04
        pad_y = (max(all_y)-min(all_y))*0.04
        ax.set_xlim(min(all_x)-pad_x, max(all_x)+pad_x)
        ax.set_ylim(min(all_y)-pad_y, max(all_y)+pad_y)

    sm = plt.cm.ScalarMappable(cmap=cmap, norm=plt.Normalize(0, vmax))
    sm.set_array([])
    cbar = fig.colorbar(sm, ax=ax, shrink=0.45, pad=0.02, aspect=25)
    cbar.set_label("Penetration % (Members / Adults 18+)", color=DGRAY, fontsize=9)
    cbar.ax.tick_params(colors=DGRAY, labelsize=8)

    fig.tight_layout()
    return fig2b64(fig)


def chart_city_bubble_map(city_df, county_geo, zip_latlng):
    """City bubble map — circle size = members, colour = penetration, top 15 labeled."""
    from matplotlib.patches import Polygon as MplPoly
    from matplotlib.collections import PatchCollection

    cmap = mcolors.LinearSegmentedColormap.from_list("pen", [RED, ORANGE, GOLD, GREEN])

    # Build city centroids from zip lat/lng
    city_coords = {}
    for z, info in zip_latlng.items():
        key = (info["city"].strip(), info["county"].strip())
        if key not in city_coords:
            city_coords[key] = []
        city_coords[key].append((info["lng"], info["lat"]))

    # Merge with city_df
    plot_cities = []
    for _, row in city_df[city_df["members"] > 500].iterrows():
        key = (row["city"].strip(), row["county"].strip())
        coords = city_coords.get(key)
        if not coords:
            # fallback: county name match
            for (c, cn), pts in city_coords.items():
                if cn == row["county"] and c == row["city"]:
                    coords = pts
                    break
        if coords:
            lng = sum(p[0] for p in coords) / len(coords)
            lat = sum(p[1] for p in coords) / len(coords)
            plot_cities.append({
                "city": row["city"], "county": row["county"],
                "members": row["members"], "pen": row["pen"],
                "adults": row["adults"],
                "lng": lng, "lat": lat
            })

    fig, ax = plt.subplots(figsize=(14, 10), facecolor=WHITE)
    ax.set_facecolor("#D6EAF8")
    ax.set_aspect("equal")
    ax.axis("off")
    ax.set_title("AAA WCNY Territory — Members by City (bubble = size, colour = penetration %)",
                 color=NAVY, fontsize=14, fontweight="bold", pad=16)

    # Draw county outlines as grey backdrop
    for name, geom in county_geo.items():
        coords = geom.get("coordinates", [])
        rings = [coords[0]] if geom["type"] == "Polygon" else [p[0] for p in coords]
        for ring in rings:
            pts = [(x, y) for x, y in ring]
            if len(pts) < 3:
                continue
            poly = MplPoly(pts, closed=True)
            pc = PatchCollection([poly], facecolor="#EBF5FB",
                                 edgecolor=MGRAY, linewidths=0.6, zorder=1)
            ax.add_collection(pc)

    if plot_cities:
        vmax = max(c["pen"] for c in plot_cities)
        max_mem = max(c["members"] for c in plot_cities)
        for c in plot_cities:
            sz = (c["members"] / max_mem) * 1200 + 20
            col = cmap(c["pen"] / vmax)
            ax.scatter(c["lng"], c["lat"], s=sz, c=[col],
                       alpha=0.75, edgecolors=WHITE, linewidths=0.6, zorder=3)

        # Label top 15 by adult population
        top15 = sorted(plot_cities, key=lambda x: x["adults"], reverse=True)[:15]
        for c in top15:
            ax.annotate(
                f"{c['city']}\n{c['pen']:.1f}%",
                (c["lng"], c["lat"]),
                xytext=(5, 5), textcoords="offset points",
                fontsize=6.5, fontweight="bold", color=NAVY, zorder=4,
                bbox=dict(boxstyle="round,pad=0.2", fc=WHITE, ec=MGRAY,
                          alpha=0.85, linewidth=0.5)
            )

        # Set extent from county geometry
        all_coords = [(x, y) for geom in county_geo.values()
                      for ring in ([geom["coordinates"][0]] if geom["type"]=="Polygon"
                                   else [p[0] for p in geom["coordinates"]])
                      for x, y in ring]
        if all_coords:
            xs = [p[0] for p in all_coords]
            ys = [p[1] for p in all_coords]
            pad_x = (max(xs)-min(xs))*0.04
            pad_y = (max(ys)-min(ys))*0.04
            ax.set_xlim(min(xs)-pad_x, max(xs)+pad_x)
            ax.set_ylim(min(ys)-pad_y, max(ys)+pad_y)

        sm = plt.cm.ScalarMappable(cmap=cmap, norm=plt.Normalize(0, vmax))
        sm.set_array([])
        cbar = fig.colorbar(sm, ax=ax, shrink=0.4, pad=0.02, aspect=25)
        cbar.set_label("Penetration %", color=DGRAY, fontsize=9)
        cbar.ax.tick_params(colors=DGRAY, labelsize=8)

    fig.tight_layout()
    return fig2b64(fig)


def build_zip_df(mem,census,dmv):
    rows=[]
    for z,m in mem.items():
        c=census.get(z,{}); d=dmv.get(z,{})
        ad=c.get("adults",0); hh=c.get("hh",0); v=d.get("vehicles",0); ov=d.get("old",0)
        rows.append(dict(zip=z,city=m["city"],county=m["county"],
            members=int(m["members"]),pop=int(m["pop"]),adults=int(ad),hh=int(hh),
            pen=round(m["members"]/ad*100,1) if ad else 0,
            vehicles=int(v),old_vehicles=int(ov),
            old_pct=round(ov/v*100,1) if v else 0,
            veh_pen=round(m["members"]/v*100,1) if v else 0,
            gap_adults=max(0,int(ad-m["members"])),
            gap_hh=max(0,int(hh-m["members"])),
            income=int(c.get("income",0)),
            median_age=c.get("age",0),college=c.get("college",0),
            new_vehicles=int(d.get("new",0))))
    return pd.DataFrame(rows)


def agg(df, keys):
    g=df.groupby(keys).agg(
        members=("members","sum"),pop=("pop","sum"),adults=("adults","sum"),
        hh=("hh","sum"),vehicles=("vehicles","sum"),old_vehicles=("old_vehicles","sum"),
        new_vehicles=("new_vehicles","sum"),
        gap_adults=("gap_adults","sum"),gap_hh=("gap_hh","sum"),zips=("zip","count")).reset_index()
    g["pen"]=(g["members"]/g["adults"].replace(0,1)*100).round(1)
    g["veh_pen"]=(g["members"]/g["vehicles"].replace(0,1)*100).round(1)
    g["old_pct"]=(g["old_vehicles"]/g["vehicles"].replace(0,1)*100).round(1)
    g["share"]=(g["members"]/TOTAL_MEMBERS*100).round(1)
    g["rev_M"]=(g["gap_adults"]*100/1e6).round(1)
    return g


def fig2b64(fig, dpi=200):
    buf=io.BytesIO()
    fig.savefig(buf,format="png",dpi=dpi,bbox_inches="tight",facecolor=fig.get_facecolor())
    plt.close(fig); buf.seek(0)
    return base64.b64encode(buf.read()).decode()


def clean_ax(ax, title="", xl="", yl=""):
    ax.set_facecolor(WHITE)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.spines["bottom"].set_color(MGRAY)
    ax.spines["left"].set_color(MGRAY)
    ax.tick_params(colors=DGRAY, labelsize=8)
    if title: ax.set_title(title, color=NAVY, fontsize=13, fontweight="bold", pad=12)
    if xl: ax.set_xlabel(xl, color=DGRAY, fontsize=9)
    if yl: ax.set_ylabel(yl, color=DGRAY, fontsize=9)


def draw_pie3d(ax, vals, labels, colors, title=""):
    ax.set_facecolor(WHITE)
    ax.set_xlim(-1.45, 1.45); ax.set_ylim(-1.12, 1.1)
    ax.set_aspect("equal"); ax.axis("off")
    if title:
        ax.set_title(title, color=NAVY, fontsize=15, fontweight="bold", pad=16)
    total = sum(vals); fracs = [v/total for v in vals]
    depth = 0.10
    angle = 90
    for i, frac in enumerate(fracs):
        t1, t2 = angle, angle + frac * 360
        rgba = to_rgba(colors[i % len(colors)])
        dark = (rgba[0]*0.6, rgba[1]*0.6, rgba[2]*0.6, 0.9)
        npts = max(int(frac * 80), 3)
        thetas = np.linspace(np.radians(t1), np.radians(t2), npts)
        for j in range(len(thetas)-1):
            x = [np.cos(thetas[j]), np.cos(thetas[j+1]),
                 np.cos(thetas[j+1]), np.cos(thetas[j])]
            y = [np.sin(thetas[j])*0.72-depth, np.sin(thetas[j+1])*0.72-depth,
                 np.sin(thetas[j+1])*0.72, np.sin(thetas[j])*0.72]
            ax.fill(x, y, color=dark, linewidth=0)
        angle = t2
    wedges, _, autotexts = ax.pie(
        fracs, autopct=lambda p: f"{p:.1f}%" if p > 4 else "",
        colors=[colors[i % len(colors)] for i in range(len(fracs))],
        startangle=90, radius=1.0,
        wedgeprops=dict(edgecolor=WHITE, linewidth=2),
        pctdistance=0.72, center=(0, 0.05))
    for t in autotexts:
        t.set_color("white"); t.set_fontsize(9); t.set_fontweight("bold")
    ax.set_ylim(-1.12, 1.08)
    return wedges


def chart_pie_county(cdf):
    top = cdf.nlargest(8, "members")
    other = TOTAL_MEMBERS - top["members"].sum()
    fig, ax = plt.subplots(figsize=(11, 7.5), facecolor=WHITE)
    vals = list(top["members"]) + [other]
    labs = list(top["county"]) + ["Other (21)"]
    wedges = draw_pie3d(ax, vals, labs, PIE, "Membership Composition by County")
    legend_labels = [f'{l}  —  {v:,.0f}  ({v/TOTAL_MEMBERS*100:.1f}%)' for l,v in zip(labs,vals)]
    ax.legend(wedges, legend_labels, loc="center left", bbox_to_anchor=(1.02, 0.5),
              fontsize=9.5, facecolor=WHITE, edgecolor=MGRAY, labelcolor=NAVY, framealpha=1)
    fig.tight_layout()
    return fig2b64(fig)


def chart_pie_city(city_df):
    top = city_df.nlargest(10, "members")
    other = TOTAL_MEMBERS - top["members"].sum()
    fig, ax = plt.subplots(figsize=(11, 7.5), facecolor=WHITE)
    vals = list(top["members"]) + [other]
    labs = [r["city"] for _,r in top.iterrows()] + ["All Others"]
    wedges = draw_pie3d(ax, vals, labs, PIE, "Membership Composition by City")
    legend_labels = [f'{l}  —  {v:,.0f}  ({v/TOTAL_MEMBERS*100:.1f}%)' for l,v in zip(labs,vals)]
    ax.legend(wedges, legend_labels, loc="center left", bbox_to_anchor=(1.02, 0.5),
              fontsize=9, facecolor=WHITE, edgecolor=MGRAY, labelcolor=NAVY, framealpha=1)
    fig.tight_layout()
    return fig2b64(fig)


def chart_opp_cities(city_df):
    top = city_df[city_df["adults"]>5000].nlargest(20, "gap_adults").sort_values("gap_adults")
    fig, ax = plt.subplots(figsize=(13, 9.5), facecolor=WHITE)
    clean_ax(ax, "Top 20 Cities by Untapped Adult Population", xl="Untapped Adults 18+")
    cmap = mcolors.LinearSegmentedColormap.from_list("c", [RED, ORANGE, GOLD, GREEN])
    norm = plt.Normalize(0, top["pen"].max())
    colors = [cmap(norm(v)) for v in top["pen"]]
    bars = ax.barh(top["city"]+" ("+top["county"]+")", top["gap_adults"],
                   color=colors, height=0.72, edgecolor="none")
    for bar, row in zip(bars, top.itertuples()):
        rev = row.gap_adults * 100 / 1e6
        ax.text(bar.get_width()+800, bar.get_y()+bar.get_height()/2,
                f"{row.gap_adults:,}  |  ${rev:.1f}M  |  {row.pen}% pen",
                va="center", color=NAVY, fontsize=8, fontweight="bold")
    ax.xaxis.set_major_formatter(mticker.FuncFormatter(lambda x,_: f"{x/1000:.0f}K"))
    sm = plt.cm.ScalarMappable(cmap=cmap, norm=norm)
    sm.set_array([])
    cbar = fig.colorbar(sm, ax=ax, pad=0.02, aspect=30, shrink=0.5)
    cbar.set_label("Current Penetration %", color=DGRAY, fontsize=9)
    cbar.ax.tick_params(colors=DGRAY, labelsize=7)
    fig.tight_layout()
    return fig2b64(fig)


def chart_veh_vs_adult(cdf):
    top = cdf[cdf["vehicles"]>10000].nlargest(18,"adults").sort_values("pen")
    fig, ax = plt.subplots(figsize=(13, 8), facecolor=WHITE)
    clean_ax(ax, "Vehicle vs Adult Penetration by County", xl="Penetration %")
    y = np.arange(len(top)); h = 0.35
    ax.barh(y + h/2, top["pen"], h, color=NAVY, alpha=0.85, label="Adult Pen (Mem/Adults)")
    ax.barh(y - h/2, top["veh_pen"], h, color=TEAL, alpha=0.85, label="Vehicle Pen (Mem/Vehicles)")
    ax.set_yticks(y)
    ax.set_yticklabels(top["county"], fontsize=8)
    ax.legend(loc="lower right", fontsize=9, facecolor=WHITE, edgecolor=MGRAY)
    fig.tight_layout()
    return fig2b64(fig)


def chart_battery(city_df):
    b = city_df[(city_df["old_pct"]>75)&(city_df["pen"]<15)&(city_df["old_vehicles"]>2000)].copy()
    b = b.nlargest(20, "old_vehicles").sort_values("old_vehicles")
    fig, ax = plt.subplots(figsize=(13, 8.5), facecolor=WHITE)
    clean_ax(ax, "Battery & Roadside Opportunity: Aging Vehicles in Low-Penetration Markets",
             xl="Vehicles 3+ Years Old")
    bars = ax.barh(b["city"]+" ("+b["county"]+")", b["old_vehicles"],
                   color=ORANGE, height=0.7, edgecolor="none", alpha=0.88)
    for bar, row in zip(bars, b.itertuples()):
        ax.text(bar.get_width()+300, bar.get_y()+bar.get_height()/2,
                f"{row.old_vehicles:,}  |  {row.old_pct:.0f}% old fleet  |  {row.pen}% membership",
                va="center", color=NAVY, fontsize=8)
    ax.xaxis.set_major_formatter(mticker.FuncFormatter(lambda x,_: f"{x/1000:.0f}K"))
    fig.tight_layout()
    return fig2b64(fig)


def chart_income(zdf):
    d = zdf[(zdf["income"]>20000)&(zdf["adults"]>2000)&(zdf["pen"]>0)].sort_values("income")
    n=len(d); q=n//5
    labels=["Under $50K","$50K–$65K","$65K–$80K","$80K–$95K","Over $95K"]
    vals=[d.iloc[i*q:(i+1)*q]["pen"].mean() for i in range(5)]
    fig, ax = plt.subplots(figsize=(10.5, 6), facecolor=WHITE)
    clean_ax(ax, "Membership Penetration by Income Bracket", yl="Avg Penetration %")
    colors=[RED, ORANGE, "#F9A825", GREEN, "#1B5E20"]
    bars = ax.bar(labels, vals, color=colors, width=0.58, edgecolor="none")
    for bar, v in zip(bars, vals):
        ax.text(bar.get_x()+bar.get_width()/2, v+0.5, f"{v:.1f}%",
                ha="center", color=NAVY, fontsize=12, fontweight="bold")
    fig.tight_layout()
    return fig2b64(fig)


def chart_quadrant_county(cdf):
    """County scatter — works well because only ~26 points."""
    data = cdf[cdf["adults"]>10000].copy()
    fig, ax = plt.subplots(figsize=(12, 8), facecolor=WHITE)
    clean_ax(ax, "Strategic Priority Matrix — Counties",
             xl="Untapped Adults 18+", yl="Current Penetration %")
    mg = data["gap_adults"].median(); mp = data["pen"].median()
    for _, r in data.iterrows():
        if r["gap_adults"]>mg and r["pen"]>mp: col=GREEN
        elif r["gap_adults"]>mg: col=RED
        elif r["pen"]>mp: col=TEAL
        else: col=ORANGE
        sz = r["members"]/data["members"].max()*450+50
        ax.scatter(r["gap_adults"], r["pen"], s=sz, c=col, alpha=0.7,
                   edgecolors=WHITE, linewidth=0.8)
        ax.annotate(r["county"], (r["gap_adults"], r["pen"]),
                    fontsize=8, color=NAVY, fontweight="bold",
                    xytext=(6,4), textcoords="offset points")
    ax.axhline(mp, color=MGRAY, ls="--", alpha=0.5)
    ax.axvline(mg, color=MGRAY, ls="--", alpha=0.5)
    ax.text(0.98,0.97,"GROW",transform=ax.transAxes,ha="right",va="top",
            color=GREEN,fontsize=11,fontweight="bold",alpha=0.5)
    ax.text(0.98,0.03,"ENTER",transform=ax.transAxes,ha="right",va="bottom",
            color=RED,fontsize=11,fontweight="bold",alpha=0.5)
    ax.text(0.02,0.97,"PROTECT",transform=ax.transAxes,ha="left",va="top",
            color=TEAL,fontsize=11,fontweight="bold",alpha=0.5)
    ax.text(0.02,0.03,"MONITOR",transform=ax.transAxes,ha="left",va="bottom",
            color=ORANGE,fontsize=11,fontweight="bold",alpha=0.5)
    ax.xaxis.set_major_formatter(mticker.FuncFormatter(lambda x,_: f"{x/1000:.0f}K"))
    fig.tight_layout()
    return fig2b64(fig)


def chart_city_strategy(city_df):
    """Replace unreadable scatter with grouped horizontal bars — top 5 per quadrant."""
    c = city_df[city_df["adults"]>=10000].copy()
    mg = c["gap_adults"].median(); mp = c["pen"].median()

    def assign_q(r):
        if r["gap_adults"]>mg and r["pen"]>mp: return "GROW — High Gap + Strong Pen"
        if r["gap_adults"]>mg: return "ENTER — High Gap + Low Pen"
        if r["pen"]>mp: return "PROTECT — Small Gap + Strong Pen"
        return "MONITOR — Small Gap + Low Pen"

    c["quadrant"] = c.apply(assign_q, axis=1)
    q_colors = {
        "GROW — High Gap + Strong Pen": GREEN,
        "ENTER — High Gap + Low Pen": RED,
        "PROTECT — Small Gap + Strong Pen": TEAL,
        "MONITOR — Small Gap + Low Pen": ORANGE,
    }
    q_order = ["ENTER — High Gap + Low Pen", "GROW — High Gap + Strong Pen",
               "PROTECT — Small Gap + Strong Pen", "MONITOR — Small Gap + Low Pen"]

    fig, axes = plt.subplots(2, 2, figsize=(18, 12), facecolor=WHITE)

    for ax, qname in zip(axes.flat, q_order):
        subset = c[c["quadrant"]==qname].nlargest(7, "gap_adults").sort_values("gap_adults")
        color = q_colors[qname]
        clean_ax(ax, qname, xl="Untapped Adults 18+")
        if len(subset) == 0:
            ax.text(0.5, 0.5, "No cities in this quadrant", transform=ax.transAxes,
                    ha="center", va="center", color=MGRAY, fontsize=11)
            continue

        bars = ax.barh(subset["city"]+" ("+subset["county"]+")", subset["gap_adults"],
                       color=color, height=0.6, alpha=0.85)
        for bar, row in zip(bars, subset.itertuples()):
            ax.text(bar.get_width()+500, bar.get_y()+bar.get_height()/2,
                    f"{row.gap_adults:,} gap  |  {row.pen}% pen  |  {row.members:,} mem",
                    va="center", color=NAVY, fontsize=8)
        ax.xaxis.set_major_formatter(mticker.FuncFormatter(lambda x,_: f"{x/1000:.0f}K"))

    fig.suptitle("City Strategic Priorities — Top Cities per Quadrant",
                 color=NAVY, fontsize=16, fontweight="bold", y=1.01)
    fig.tight_layout()
    return fig2b64(fig)


def chart_heatmap(cdf):
    """County performance heatmap — multi-metric view in one chart."""
    top = cdf[cdf["adults"]>5000].sort_values("pen", ascending=False).copy().head(25)

    # Compute mem_per_hh
    top["mem_hh"] = np.where(top["hh"]>0, (top["members"]/top["hh"]).round(2), 0)
    cols = ["pen", "veh_pen", "old_pct", "mem_hh"]
    col_labels = ["Penetration\n(Mem/Adults)", "Vehicle Pen\n(Mem/Vehs)", "Old Fleet %", "Members\nper HH"]

    data = top[cols].values.astype(float)

    # Normalize each column 0-1 for coloring
    normed = np.zeros_like(data)
    for j in range(data.shape[1]):
        lo, hi = data[:,j].min(), data[:,j].max()
        normed[:,j] = (data[:,j]-lo)/(hi-lo) if hi>lo else 0.5

    # For old_fleet, INVERT: higher old = worse = red; lower = green
    normed[:,2] = 1 - normed[:,2]

    fig, ax = plt.subplots(figsize=(10, 10), facecolor=WHITE)
    ax.set_facecolor(WHITE)

    cmap = mcolors.LinearSegmentedColormap.from_list("hm",
        ["#FFCDD2", "#EF9A9A", "#E57373", "#EF5350", "#D32F2F",
         "#C62828", "#B71C1C"])
    # Actually use a green-to-red: good=green, bad=red
    cmap = mcolors.LinearSegmentedColormap.from_list("hm",
        [RED, "#FF8A65", "#FFE082", "#AED581", GREEN])

    im = ax.imshow(normed, cmap=cmap, aspect="auto", vmin=0, vmax=1)

    ax.set_xticks(range(len(col_labels)))
    ax.set_xticklabels(col_labels, fontsize=10, fontweight="bold", color=NAVY)
    ax.set_yticks(range(len(top)))
    ax.set_yticklabels(top["county"], fontsize=9, color=NAVY)
    ax.tick_params(length=0)

    for i in range(len(top)):
        for j in range(len(cols)):
            val = data[i,j]
            fmt = f"{val:.1f}%" if j < 3 else f"{val:.2f}"
            color = WHITE if normed[i,j] > 0.7 or normed[i,j] < 0.3 else NAVY
            ax.text(j, i, fmt, ha="center", va="center", fontsize=8.5,
                    color=color, fontweight="bold")

    ax.set_title("County Performance Heatmap", fontsize=14, fontweight="bold",
                 color=NAVY, pad=14, loc="left")
    ax.spines[:].set_visible(False)

    # Add colorbar
    cbar = fig.colorbar(im, ax=ax, shrink=0.4, pad=0.02)
    cbar.set_label("Performance (green = strong, red = weak)", color=DGRAY, fontsize=9)
    cbar.ax.tick_params(labelsize=0, length=0)

    fig.tight_layout()
    return fig2b64(fig)


def chart_hh(cdf):
    top=cdf.nlargest(15,"gap_hh").sort_values("gap_hh")
    fig, ax = plt.subplots(figsize=(12, 7.5), facecolor=WHITE)
    clean_ax(ax,"Households Without AAA Membership", xl="Non-Member Households")
    bars=ax.barh(top["county"], top["gap_hh"], color=NAVY, height=0.65, alpha=0.85)
    for bar, row in zip(bars, top.itertuples()):
        pct=row.gap_hh/row.hh*100 if row.hh else 0
        ax.text(bar.get_width()+500, bar.get_y()+bar.get_height()/2,
                f"{row.gap_hh:,}  ({pct:.0f}% of households)", va="center", color=DGRAY, fontsize=9)
    ax.xaxis.set_major_formatter(mticker.FuncFormatter(lambda x,_: f"{x/1000:.0f}K"))
    fig.tight_layout()
    return fig2b64(fig)


def chart_veh_age(cdf):
    top=cdf[cdf["vehicles"]>10000].nlargest(18,"vehicles").sort_values("vehicles")
    fig, ax = plt.subplots(figsize=(12, 7), facecolor=WHITE)
    clean_ax(ax,"Vehicle Fleet Age Distribution by County", xl="Number of Vehicles")
    ax.barh(top["county"], top["old_vehicles"], color=ORANGE, height=0.6, label="3+ Years Old", alpha=0.85)
    ax.barh(top["county"], top["vehicles"]-top["old_vehicles"], left=top["old_vehicles"],
            color=TEAL, height=0.6, label="Newer (<3 yrs)", alpha=0.85)
    ax.xaxis.set_major_formatter(mticker.FuncFormatter(lambda x,_: f"{x/1000:.0f}K"))
    ax.legend(fontsize=9, facecolor=WHITE, edgecolor=MGRAY)
    fig.tight_layout()
    return fig2b64(fig)


def chart_demo(zdf):
    d=zdf[(zdf["income"]>10000)&(zdf["adults"]>2000)&(zdf["pen"]>0)].copy()
    fig, axes = plt.subplots(1, 3, figsize=(16.5, 5.5), facecolor=WHITE)
    plots=[("income","Household Income ($K)",BLUE,1000),
           ("college","College-Educated %",PURPLE,1),
           ("median_age","Median Age",TEAL,1)]
    for ax,(col,label,color,div) in zip(axes, plots):
        clean_ax(ax, yl="Penetration %")
        xv=d[col]/div if div>1 else d[col]
        ax.scatter(xv, d["pen"], s=d["adults"]/d["adults"].max()*70,
                   c=color, alpha=0.25, edgecolors="none")
        z_fit=np.polyfit(xv, d["pen"], 1)
        xl=np.linspace(xv.min(), xv.max(), 100)
        ax.plot(xl, np.poly1d(z_fit)(xl), color=RED, lw=2, ls="--", alpha=0.7)
        corr=xv.corr(d["pen"])
        ax.set_title(f"{label}  (r = {corr:.2f})", color=NAVY, fontsize=11, fontweight="bold")
    fig.suptitle("What Drives Membership Penetration?", color=NAVY, fontsize=14,
                 fontweight="bold", y=1.01)
    fig.tight_layout()
    return fig2b64(fig)


def chart_overunder(zdf):
    d=zdf[(zdf["income"]>20000)&(zdf["adults"]>3000)&(zdf["pen"]>0)].copy()
    coef=np.polyfit(d["income"], d["pen"], 1)
    d["expected"]=np.polyval(coef, d["income"])
    d["residual"]=d["pen"]-d["expected"]
    over=d.nlargest(12, "residual").sort_values("residual")
    under=d.nsmallest(12, "residual").sort_values("residual", ascending=False)
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(18, 7), facecolor=WHITE)
    clean_ax(ax1, "Outperformers (beating income prediction)")
    bars1=ax1.barh(over["city"]+" ("+over["county"]+")", over["residual"],
                   color=GREEN, height=0.65, alpha=0.85)
    for bar, row in zip(bars1, over.itertuples()):
        ax1.text(bar.get_width()+0.3, bar.get_y()+bar.get_height()/2,
                 f"+{row.residual:.1f}pts  |  {row.pen:.0f}% actual  |  ${row.income:,}",
                 va="center", color=NAVY, fontsize=8)
    clean_ax(ax2, "Underperformers (trailing income prediction)")
    bars2=ax2.barh(under["city"]+" ("+under["county"]+")", under["residual"].abs(),
                   color=RED, height=0.65, alpha=0.85)
    for bar, row in zip(bars2, under.itertuples()):
        ax2.text(bar.get_width()+0.3, bar.get_y()+bar.get_height()/2,
                 f"−{abs(row.residual):.1f}pts  |  {row.pen:.1f}% actual  |  ${row.income:,}",
                 va="center", color=NAVY, fontsize=8)
    fig.tight_layout()
    return fig2b64(fig)


def chart_donuts(zdf):
    weak = zdf[(zdf["county"].isin(["Erie","Monroe","Onondaga","Niagara","Ontario"]))
               &(zdf["adults"]>3000)&(zdf["pen"]<15)].copy()
    weak = weak.nlargest(15, "adults").sort_values("adults")
    fig, ax = plt.subplots(figsize=(13, 7.5), facecolor=WHITE)
    clean_ax(ax, "Urban Donut Holes — Weak Spots in Our Strongest Counties", xl="Adults 18+")
    bars=ax.barh(weak["zip"]+" "+weak["city"]+" ("+weak["county"]+")",
                 weak["adults"], color=LGRAY, height=0.65, edgecolor=MGRAY, linewidth=0.5)
    ax.barh(weak["zip"]+" "+weak["city"]+" ("+weak["county"]+")",
            weak["members"], color=NAVY, height=0.65, alpha=0.85)
    for bar, row in zip(bars, weak.itertuples()):
        ax.text(bar.get_width()+200, bar.get_y()+bar.get_height()/2,
                f"{row.pen:.1f}% pen  |  {row.members:,} / {row.adults:,}  |  ${row.income:,} income",
                va="center", color=DGRAY, fontsize=8)
    ax.legend(handles=[Patch(facecolor=NAVY, label="Current Members"),
                       Patch(facecolor=LGRAY, edgecolor=MGRAY, label="Total Adults")],
              fontsize=9, facecolor=WHITE, edgecolor=MGRAY)
    ax.xaxis.set_major_formatter(mticker.FuncFormatter(lambda x,_: f"{x/1000:.0f}K"))
    fig.tight_layout()
    return fig2b64(fig)


def chart_vehicle_geo_map(county_df, county_geo):
    """Side-by-side county choropleth: old fleet % (left) and new fleet % (right)."""
    from matplotlib.patches import Polygon as MplPoly
    from matplotlib.collections import PatchCollection

    old_veh_map = dict(zip(county_df["county"], county_df["old_vehicles"]))
    new_veh_map = dict(zip(county_df["county"], county_df["new_vehicles"]))

    configs = [
        (old_veh_map, "Aging Fleet — Vehicles 3+ Years Old (count)",
         [TEAL, GOLD, ORANGE, RED],  "Vehicles (old)"),
        (new_veh_map, "Modern Fleet — Vehicles Under 3 Years Old (count)",
         [RED,  GOLD, TEAL,  GREEN], "Vehicles (new)"),
    ]

    fig, axes = plt.subplots(1, 2, figsize=(18, 9), facecolor=WHITE)
    fig.suptitle("Vehicle Fleet Age — County Geographic View",
                 color=NAVY, fontsize=14, fontweight="bold")

    for ax, (data_map, title, colors, label) in zip(axes, configs):
        ax.set_facecolor("#D6EAF8")
        ax.set_aspect("equal")
        ax.axis("off")
        ax.set_title(title, color=NAVY, fontsize=11, fontweight="bold", pad=12)

        vals = [v for v in data_map.values() if v > 0]
        vmin, vmax = (min(vals), max(vals)) if vals else (0, 1)
        cmap = mcolors.LinearSegmentedColormap.from_list("v", colors)
        norm = plt.Normalize(vmin, vmax)

        plotted = []
        for name, geom in county_geo.items():
            val = data_map.get(name, 0)
            rings = ([geom["coordinates"][0]] if geom["type"] == "Polygon"
                     else [p[0] for p in geom["coordinates"]])
            for ring in rings:
                pts = [(x, y) for x, y in ring]
                if len(pts) < 3:
                    continue
                pc = PatchCollection([MplPoly(pts, closed=True)],
                                     facecolor=cmap(norm(val)),
                                     edgecolor=WHITE, linewidths=0.7, zorder=2)
                ax.add_collection(pc)
                plotted.append(pts)

            if rings:
                xs = [p[0] for p in rings[0]]
                ys = [p[1] for p in rings[0]]
                cx, cy = sum(xs)/len(xs), sum(ys)/len(ys)
                label_val = f"{val/1000:.0f}K" if val >= 1000 else str(int(val))
                ax.text(cx, cy, f"{name}\n{label_val}",
                        ha="center", va="center", fontsize=6,
                        color=WHITE if norm(val) > 0.65 else NAVY,
                        fontweight="bold", zorder=3)

        if plotted:
            all_x = [p[0] for pts in plotted for p in pts]
            all_y = [p[1] for pts in plotted for p in pts]
            px = (max(all_x)-min(all_x))*0.04; py = (max(all_y)-min(all_y))*0.04
            ax.set_xlim(min(all_x)-px, max(all_x)+px)
            ax.set_ylim(min(all_y)-py, max(all_y)+py)

        sm = plt.cm.ScalarMappable(cmap=cmap, norm=norm)
        sm.set_array([])
        cbar = fig.colorbar(sm, ax=ax, shrink=0.4, pad=0.02, aspect=25)
        cbar.set_label(label, color=DGRAY, fontsize=8)
        cbar.ax.yaxis.set_major_formatter(
            mticker.FuncFormatter(lambda x, _: f"{x/1000:.0f}K" if x >= 1000 else f"{int(x)}"))
        cbar.ax.tick_params(colors=DGRAY, labelsize=7)

    fig.tight_layout()
    return fig2b64(fig)


# ══════════════════════════════════════════════════════════════════════════
# HTML — exact v4 structure with Big Picture narrative
# ══════════════════════════════════════════════════════════════════════════
def build(zdf, city_df, county_df, charts):
    today = datetime.now().strftime("%B %d, %Y")
    T=TOTAL_MEMBERS; ad=zdf["adults"].sum(); hh=zdf["hh"].sum()
    pop=zdf["pop"].sum(); veh=zdf["vehicles"].sum(); ov=zdf["old_vehicles"].sum()
    pen=T/ad*100; gap=ad-T
    top5=county_df.head(5)["share"].sum()
    crit=county_df[county_df["pen"]<5]

    def sec(n,t,q=""):
        return f"""<div style="margin:45px 0 18px;page-break-before:always;">
        <div style="border-bottom:3px solid {NAVY};padding-bottom:8px;">
        <span style="color:{TEAL};font-size:12px;font-weight:700;letter-spacing:1px;">SECTION {n}</span>
        <div style="color:{NAVY};font-size:20px;font-weight:700;margin-top:4px;">{t}</div>
        {"<div style='color:"+DGRAY+";font-size:11px;font-style:italic;margin-top:2px;'>"+q+"</div>" if q else ""}
        </div></div>"""

    def ins(text):
        return f"""<div style="background:{LGRAY};border-left:4px solid {TEAL};
        padding:14px 18px;margin:14px 0;font-size:12px;line-height:1.75;color:{DARK};">{text}</div>"""

    def rec(p,t,target,bullets,impact,c):
        bl="".join(f"<li>{b}</li>" for b in bullets)
        return f"""<div style="background:{LGRAY};border-left:4px solid {c};border-radius:0 6px 6px 0;
        padding:16px 18px;margin:12px 0;">
        <div style="color:{c};font-size:10px;font-weight:700;letter-spacing:1px;margin-bottom:4px;">PRIORITY {p}</div>
        <div style="color:{NAVY};font-size:14px;font-weight:700;margin-bottom:6px;">{t}</div>
        <div style="color:{DGRAY};font-size:11px;margin-bottom:6px;"><strong>Target:</strong> {target}</div>
        <ul style="color:{DGRAY};font-size:11px;line-height:1.8;padding-left:16px;">{bl}</ul>
        <div style="margin-top:8px;padding:6px 12px;background:{c}12;border-radius:4px;
        font-size:11px;color:{c};font-weight:600;">Expected Impact: {impact}</div></div>"""

    html = f"""<!DOCTYPE html><html><head><meta charset="utf-8">
<title>AAA WCNY — Membership Strategic Analysis</title>
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap');
*{{margin:0;padding:0;box-sizing:border-box;}}
body{{font-family:'Inter',sans-serif;background:{WHITE};color:{DARK};line-height:1.65;
font-size:12.5px;-webkit-print-color-adjust:exact;print-color-adjust:exact;}}
@media print{{.pb{{page-break-before:always;}}}}
.c{{max-width:1060px;margin:0 auto;padding:40px 55px;}}
.img{{width:100%;border-radius:4px;margin:12px 0;border:1px solid #E0E0E0;}}
table{{width:100%;border-collapse:collapse;font-size:10.5px;margin:12px 0;}}
th{{background:{NAVY};color:{WHITE};padding:9px 8px;text-align:left;font-size:10px;font-weight:600;}}
td{{padding:8px;border-bottom:1px solid #E8EBF0;color:{DARK};}}
tr:nth-child(even){{background:{LGRAY};}}
h2{{color:{NAVY};font-size:18px;font-weight:700;margin:30px 0 12px;}}
p{{margin:8px 0;}}
</style></head><body><div class="c">

<!-- COVER -->
<div style="min-height:88vh;display:flex;flex-direction:column;justify-content:center;">
<div style="border-left:5px solid {TEAL};padding-left:25px;">
<div style="font-size:12px;letter-spacing:4px;color:{DGRAY};text-transform:uppercase;">AAA Western & Central New York</div>
<div style="font-size:36px;font-weight:800;color:{NAVY};line-height:1.15;margin:15px 0;">
Membership Penetration<br>Strategic Analysis</div>
<div style="font-size:13px;color:{DGRAY};margin-top:12px;">
Prepared for the Board of Directors, CEO, and Head of Membership<br>
{today} &nbsp;&bull;&nbsp; Confidential</div>
</div>
<div style="display:flex;gap:30px;margin-top:50px;">
<div style="text-align:center;">
<div style="color:{NAVY};font-size:42px;font-weight:800;">{T:,}</div>
<div style="color:{DGRAY};font-size:11px;">Total Members</div></div>
<div style="text-align:center;">
<div style="color:{TEAL};font-size:42px;font-weight:800;">{pen:.1f}%</div>
<div style="color:{DGRAY};font-size:11px;">Penetration (Adults 18+)</div></div>
<div style="text-align:center;">
<div style="color:{RED};font-size:42px;font-weight:800;">${gap*100/1e6:.0f}M</div>
<div style="color:{DGRAY};font-size:11px;">Untapped Annual Revenue</div></div>
</div></div>

<!-- TABLE OF CONTENTS -->
<div style="page-break-before:always;">
<div style="margin-bottom:32px;">
<div style="font-size:11px;font-weight:700;letter-spacing:2px;color:{TEAL};text-transform:uppercase;margin-bottom:6px;">AAA Western &amp; Central New York</div>
<div style="color:{NAVY};font-size:28px;font-weight:800;line-height:1.1;">Table of Contents</div>
<div style="width:48px;height:3px;background:{TEAL};margin-top:12px;border-radius:2px;"></div>
</div>

<div style="display:flex;flex-direction:column;gap:2px;">

<div style="display:flex;align-items:center;padding:12px 16px;border-radius:6px;border:1px solid #E8EBF0;">
<div style="width:44px;min-width:44px;font-size:11px;font-weight:800;color:{TEAL};letter-spacing:1px;">EXEC</div>
<div style="flex:1;">
<div style="font-size:13px;font-weight:700;color:{NAVY};">Executive Summary — The Big Picture</div>
<div style="font-size:10.5px;color:{DGRAY};margin-top:1px;">6 key takeaways &middot; hero KPIs &middot; path to 1 million members</div>
</div>
</div>

<div style="display:flex;align-items:center;padding:12px 16px;border-radius:6px;background:{LGRAY};border:1px solid #E8EBF0;">
<div style="width:44px;min-width:44px;font-size:11px;font-weight:800;color:{TEAL};letter-spacing:1px;">GEO</div>
<div style="flex:1;">
<div style="font-size:13px;font-weight:700;color:{NAVY};">Geographic Overview — Territory Maps</div>
<div style="font-size:10.5px;color:{DGRAY};margin-top:1px;">County penetration choropleth &middot; city member bubble map</div>
</div>
</div>

<div style="display:flex;align-items:center;padding:12px 16px;border-radius:6px;border:1px solid #E8EBF0;">
<div style="width:44px;min-width:44px;font-size:13px;font-weight:800;color:{TEAL};">01</div>
<div style="flex:1;">
<div style="font-size:13px;font-weight:700;color:{NAVY};">Business Composition</div>
<div style="font-size:10.5px;color:{DGRAY};margin-top:1px;">Where does our membership come from? County and city breakdown.</div>
</div>
</div>

<div style="display:flex;align-items:center;padding:12px 16px;border-radius:6px;background:{LGRAY};border:1px solid #E8EBF0;">
<div style="width:44px;min-width:44px;font-size:13px;font-weight:800;color:{TEAL};">02</div>
<div style="flex:1;">
<div style="font-size:13px;font-weight:700;color:{NAVY};">Penetration — Best &amp; Worst Markets</div>
<div style="font-size:10.5px;color:{DGRAY};margin-top:1px;">Top 12 and bottom 12 cities ranked by member penetration rate.</div>
</div>
</div>

<div style="display:flex;align-items:center;padding:12px 16px;border-radius:6px;border:1px solid #E8EBF0;">
<div style="width:44px;min-width:44px;font-size:13px;font-weight:800;color:{TEAL};">03</div>
<div style="flex:1;">
<div style="font-size:13px;font-weight:700;color:{NAVY};">Top 20 Opportunity Cities</div>
<div style="font-size:10.5px;color:{DGRAY};margin-top:1px;">Largest untapped adult populations and revenue opportunity by city.</div>
</div>
</div>

<div style="display:flex;align-items:center;padding:12px 16px;border-radius:6px;background:{LGRAY};border:1px solid #E8EBF0;">
<div style="width:44px;min-width:44px;font-size:13px;font-weight:800;color:{TEAL};">04</div>
<div style="flex:1;">
<div style="font-size:13px;font-weight:700;color:{NAVY};">Hidden Intelligence</div>
<div style="font-size:10.5px;color:{DGRAY};margin-top:1px;">Outperformers vs. underperformers &middot; urban donut holes in strong counties.</div>
</div>
</div>

<div style="display:flex;align-items:center;padding:12px 16px;border-radius:6px;border:1px solid #E8EBF0;">
<div style="width:44px;min-width:44px;font-size:13px;font-weight:800;color:{TEAL};">05</div>
<div style="flex:1;">
<div style="font-size:13px;font-weight:700;color:{NAVY};">Household Eligibility Gap</div>
<div style="font-size:10.5px;color:{DGRAY};margin-top:1px;">How many households are eligible but still without AAA membership?</div>
</div>
</div>

<div style="display:flex;align-items:center;padding:12px 16px;border-radius:6px;background:{LGRAY};border:1px solid #E8EBF0;">
<div style="width:44px;min-width:44px;font-size:13px;font-weight:800;color:{TEAL};">06</div>
<div style="flex:1;">
<div style="font-size:13px;font-weight:700;color:{NAVY};">Vehicle &amp; Fleet Intelligence</div>
<div style="font-size:10.5px;color:{DGRAY};margin-top:1px;">Vehicle penetration &middot; aging fleet by county &middot; geographic fleet maps.</div>
</div>
</div>

<div style="display:flex;align-items:center;padding:12px 16px;border-radius:6px;border:1px solid #E8EBF0;">
<div style="width:44px;min-width:44px;font-size:13px;font-weight:800;color:{TEAL};">07</div>
<div style="flex:1;">
<div style="font-size:13px;font-weight:700;color:{NAVY};">Battery &amp; Roadside Opportunity</div>
<div style="font-size:10.5px;color:{DGRAY};margin-top:1px;">409K aging vehicles in low-penetration markets — $6.1M service opportunity.</div>
</div>
</div>

<div style="display:flex;align-items:center;padding:12px 16px;border-radius:6px;background:{LGRAY};border:1px solid #E8EBF0;">
<div style="width:44px;min-width:44px;font-size:13px;font-weight:800;color:{TEAL};">08</div>
<div style="flex:1;">
<div style="font-size:13px;font-weight:700;color:{NAVY};">What Drives Membership?</div>
<div style="font-size:10.5px;color:{DGRAY};margin-top:1px;">Income, education, and age as predictors &middot; demographic correlations.</div>
</div>
</div>

<div style="display:flex;align-items:center;padding:12px 16px;border-radius:6px;border:1px solid #E8EBF0;">
<div style="width:44px;min-width:44px;font-size:13px;font-weight:800;color:{TEAL};">09</div>
<div style="flex:1;">
<div style="font-size:13px;font-weight:700;color:{NAVY};">County Performance Heatmap</div>
<div style="font-size:10.5px;color:{DGRAY};margin-top:1px;">All 29 counties across penetration, vehicle coverage, and fleet age.</div>
</div>
</div>

<div style="display:flex;align-items:center;padding:12px 16px;border-radius:6px;background:{LGRAY};border:1px solid #E8EBF0;">
<div style="width:44px;min-width:44px;font-size:13px;font-weight:800;color:{TEAL};">10</div>
<div style="flex:1;">
<div style="font-size:13px;font-weight:700;color:{NAVY};">Strategic Priority Matrix</div>
<div style="font-size:10.5px;color:{DGRAY};margin-top:1px;">ENTER · GROW · PROTECT · MONITOR — where to invest at county and city level.</div>
</div>
</div>

<div style="display:flex;align-items:center;padding:12px 16px;border-radius:6px;border:1px solid #E8EBF0;">
<div style="width:44px;min-width:44px;font-size:13px;font-weight:800;color:{TEAL};">11</div>
<div style="flex:1;">
<div style="font-size:13px;font-weight:700;color:{NAVY};">Market Spotlight: Ithaca / Tompkins County</div>
<div style="font-size:10.5px;color:{DGRAY};margin-top:1px;">Cornell student pipeline &middot; education-driven penetration &middot; replication model.</div>
</div>
</div>

<div style="display:flex;align-items:center;padding:12px 16px;border-radius:6px;background:{LGRAY};border:1px solid #E8EBF0;">
<div style="width:44px;min-width:44px;font-size:13px;font-weight:800;color:{TEAL};">12</div>
<div style="flex:1;">
<div style="font-size:13px;font-weight:700;color:{NAVY};">Strategic Recommendations</div>
<div style="font-size:10.5px;color:{DGRAY};margin-top:1px;">5 priority initiatives — projected +210,000 members &middot; +$29.6M annual revenue.</div>
</div>
</div>

<div style="display:flex;align-items:center;padding:12px 16px;border-radius:6px;border:1px solid #E8EBF0;border-top:2px solid {MGRAY};margin-top:6px;">
<div style="width:44px;min-width:44px;font-size:11px;font-weight:700;color:{MGRAY};">APP</div>
<div style="flex:1;">
<div style="font-size:12px;font-weight:600;color:{DGRAY};">Data Sources &amp; Methodology</div>
<div style="font-size:10.5px;color:{MGRAY};margin-top:1px;">AAA WCNY records &middot; U.S. Census Bureau &middot; NY State DMV registrations</div>
</div>
</div>

</div>
</div>

<!-- EXECUTIVE SUMMARY — THE BIG PICTURE -->
<div style="page-break-before:always;">
<div style="border-bottom:3px solid {NAVY};padding-bottom:8px;margin-bottom:20px;">
<span style="color:{TEAL};font-size:12px;font-weight:700;letter-spacing:1px;">EXECUTIVE SUMMARY</span>
<div style="color:{NAVY};font-size:20px;font-weight:700;margin-top:4px;">The Big Picture</div>
</div>

<!-- KEY TAKEAWAYS -->
<div style="border:1px solid #E8EBF0;border-radius:8px;overflow:hidden;margin:0 0 22px;">
<div style="background:{LGRAY};padding:12px 20px;border-bottom:1px solid #E8EBF0;">
<span style="font-size:10px;font-weight:700;letter-spacing:2px;text-transform:uppercase;color:{TEAL};">Key Takeaways</span>
</div>
<div style="display:flex;flex-direction:column;">
<div style="display:flex;align-items:baseline;gap:16px;padding:13px 20px;border-bottom:1px solid #F0F2F5;">
<span style="font-size:18px;font-weight:800;color:{TEAL};min-width:24px;line-height:1;">1</span>
<p style="margin:0;font-size:12px;line-height:1.6;color:{DARK};"><strong style="color:{NAVY};">{T:,} members, {pen:.1f}% penetration.</strong> Three in four eligible adults are not members. At $100/year, the unconverted {gap:,} adults represent <strong style="color:{RED};">${gap*100/1e6:.0f}M in unrealized annual revenue.</strong></p>
</div>
<div style="display:flex;align-items:baseline;gap:16px;padding:13px 20px;border-bottom:1px solid #F0F2F5;background:{LGRAY};">
<span style="font-size:18px;font-weight:800;color:{TEAL};min-width:24px;line-height:1;">2</span>
<p style="margin:0;font-size:12px;line-height:1.6;color:{DARK};"><strong style="color:{NAVY};">Five counties hold {top5:.0f}% of all members.</strong> Erie and Monroe alone are half the book. The remaining 24 counties — including six below 5% penetration — are nearly untouched.</p>
</div>
<div style="display:flex;align-items:baseline;gap:16px;padding:13px 20px;border-bottom:1px solid #F0F2F5;">
<span style="font-size:18px;font-weight:800;color:{RED};min-width:24px;line-height:1;">3</span>
<p style="margin:0;font-size:12px;line-height:1.6;color:{DARK};"><strong style="color:{NAVY};">Dark counties are not low-demand.</strong> Marcy (Oneida Co.) has $99K median income — 1.4% penetration. Binghamton: 0.2%. This is a <strong>distribution and awareness failure</strong>, not a product-market fit problem.</p>
</div>
<div style="display:flex;align-items:baseline;gap:16px;padding:13px 20px;border-bottom:1px solid #F0F2F5;background:{LGRAY};">
<span style="font-size:18px;font-weight:800;color:{ORANGE};min-width:24px;line-height:1;">4</span>
<p style="margin:0;font-size:12px;line-height:1.6;color:{DARK};"><strong style="color:{NAVY};">{veh:,} registered vehicles — {veh//T:.0f} per member.</strong> <strong>{ov/veh*100:.0f}% are 3+ years old</strong> — prime roadside and battery candidates. In dark counties, aging fleet rates reach 85–90%.</p>
</div>
<div style="display:flex;align-items:baseline;gap:16px;padding:13px 20px;border-bottom:1px solid #F0F2F5;">
<span style="font-size:18px;font-weight:800;color:{TEAL};min-width:24px;line-height:1;">5</span>
<p style="margin:0;font-size:12px;line-height:1.6;color:{DARK};"><strong style="color:{NAVY};">Income is the #1 predictor</strong> (r = 0.55): markets above $95K penetrate at 32.7% vs 11.3% below $50K. But Churchville, Lewiston, and Springville beat their income expectation by 17–21 pts — local execution matters.</p>
</div>
<div style="display:flex;align-items:baseline;gap:16px;padding:13px 20px;background:#F0FFF4;">
<span style="font-size:18px;font-weight:800;color:{GREEN};min-width:24px;line-height:1;">6</span>
<p style="margin:0;font-size:12px;line-height:1.6;color:{DARK};">Five targeted initiatives — dark counties, aging fleets, urban donut holes, students, essentials tier — are projected to add <strong style="color:{GREEN};">+210,000 members</strong> and <strong style="color:{GREEN};">+$29.6M annual revenue.</strong></p>
</div>
</div>
</div>

<p>AAA WCNY serves <strong>843,000 members</strong> across a 29-county territory with <strong>{ad:,} eligible adults</strong>, a penetration rate of <strong>{pen:.1f}%</strong>. While this represents a significant footprint, it also means that <strong>more than three out of four eligible adults in our territory are not members</strong>. At an average membership value of $100 per year, the untapped {gap:,} adults represent approximately <strong>${gap*100/1e6:.0f} million in unrealized annual revenue</strong>.</p>

<p>Our membership is heavily concentrated: <strong>five counties — Erie, Monroe, Onondaga, Niagara, and Ontario — account for {top5:.0f}% of all members</strong>. Even in these strongholds, penetration ranges from only 24% to 38%, meaning the majority of eligible adults remain unconverted. The remaining 24 counties split the other {100-top5:.0f}%.</p>

<p>The most striking finding is the existence of <strong>six "dark" counties</strong> — Herkimer, Tioga, Chemung, Oneida, Orleans, and Chautauqua — where penetration falls below 5% despite a combined eligible adult population of over {crit['adults'].sum():,}. These are not low-demand markets. Marcy in Oneida County, for example, has a median household income of $99,000 but a penetration rate of just 1.4%. <strong>This is a distribution and awareness problem, not a product-market fit problem.</strong></p>

<p>Vehicle data reinforces the opportunity: <strong>{veh:,} vehicles</strong> are registered in the territory — {veh//T:.0f} for every member. <strong>{ov/veh*100:.0f}% of these vehicles are three or more years old</strong>, the exact segment most likely to need roadside assistance and battery service. In the dark counties, aging fleet percentages rise to 85-90%. Every one of these vehicles represents a potential membership sale and service revenue opportunity.</p>

<p>Income is the strongest predictor of membership: markets above $95,000 median income show 32.7% penetration versus 11.3% for those below $50,000. However, several mid-income markets dramatically outperform their demographic prediction — Churchville, Lewiston, and Springville exceed expected penetration by 17-21 percentage points. <strong>Something about local execution in these communities is working, and it should be studied and replicated.</strong></p>

<p style="background:{LGRAY};padding:14px 18px;border-radius:4px;margin-top:16px;">
<strong style="color:{NAVY};">In summary:</strong> AAA WCNY has a clear path from 843,000 to over one million members. The growth will come from three sources: activating dark counties where we are virtually invisible, closing urban donut holes within our strongest markets, and leveraging the aging vehicle fleet as both a service revenue line and a membership acquisition channel. The detailed analysis and recommendations that follow lay out the evidence and the plan.
</p>
</div>

<!-- KEY METRICS -->
<div style="display:flex;flex-wrap:wrap;gap:16px;margin:30px 0;">
<div style="flex:1;min-width:140px;padding:16px;border:1px solid #E0E0E0;border-radius:6px;text-align:center;">
<div style="color:{DGRAY};font-size:9px;text-transform:uppercase;letter-spacing:1.5px;">Population</div>
<div style="color:{NAVY};font-size:22px;font-weight:800;">{pop:,}</div></div>
<div style="flex:1;min-width:140px;padding:16px;border:1px solid #E0E0E0;border-radius:6px;text-align:center;">
<div style="color:{DGRAY};font-size:9px;text-transform:uppercase;letter-spacing:1.5px;">Adults 18+</div>
<div style="color:{NAVY};font-size:22px;font-weight:800;">{ad:,}</div></div>
<div style="flex:1;min-width:140px;padding:16px;border:1px solid #E0E0E0;border-radius:6px;text-align:center;">
<div style="color:{DGRAY};font-size:9px;text-transform:uppercase;letter-spacing:1.5px;">Households</div>
<div style="color:{NAVY};font-size:22px;font-weight:800;">{hh:,}</div></div>
<div style="flex:1;min-width:140px;padding:16px;border:1px solid #E0E0E0;border-radius:6px;text-align:center;">
<div style="color:{DGRAY};font-size:9px;text-transform:uppercase;letter-spacing:1.5px;">Vehicles (DMV)</div>
<div style="color:{NAVY};font-size:22px;font-weight:800;">{veh:,}</div></div>
<div style="flex:1;min-width:140px;padding:16px;border:1px solid #E0E0E0;border-radius:6px;text-align:center;">
<div style="color:{DGRAY};font-size:9px;text-transform:uppercase;letter-spacing:1.5px;">Penetration</div>
<div style="color:{TEAL};font-size:22px;font-weight:800;">{pen:.1f}%</div></div>
<div style="flex:1;min-width:140px;padding:16px;border:1px solid #E0E0E0;border-radius:6px;text-align:center;">
<div style="color:{DGRAY};font-size:9px;text-transform:uppercase;letter-spacing:1.5px;">Revenue Gap</div>
<div style="color:{RED};font-size:22px;font-weight:800;">${gap*100/1e6:.0f}M/yr</div></div>
</div>

<!-- SECTIONS -->
<div style="margin:45px 0 18px;page-break-before:always;">
<div style="border-bottom:3px solid {NAVY};padding-bottom:8px;">
<span style="color:{TEAL};font-size:12px;font-weight:700;letter-spacing:1px;">GEOGRAPHIC OVERVIEW</span>
<div style="color:{NAVY};font-size:20px;font-weight:700;margin-top:4px;">County Penetration Map</div>
<div style="color:{DGRAY};font-size:11px;font-style:italic;margin-top:2px;">Where are we strong — and where are we invisible?</div>
</div></div>
<img class="img" src="data:image/png;base64,{charts['county_map']}">
{ins("Red counties are near-zero penetration — not low demand, but zero AAA presence. Green counties are our core. Every red county on this map is a market-entry opportunity.")}

<div style="margin:45px 0 18px;page-break-before:always;">
<div style="border-bottom:3px solid {NAVY};padding-bottom:8px;">
<span style="color:{TEAL};font-size:12px;font-weight:700;letter-spacing:1px;">GEOGRAPHIC OVERVIEW</span>
<div style="color:{NAVY};font-size:20px;font-weight:700;margin-top:4px;">City Member Distribution Map</div>
<div style="color:{DGRAY};font-size:11px;font-style:italic;margin-top:2px;">Bubble size = member count · Color = penetration % · Top 15 cities labeled</div>
</div></div>
<img class="img" src="data:image/png;base64,{charts['city_map']}">
{ins("Large red bubbles are the biggest prize — high member counts but low penetration. Small green bubbles are saturated. The white space between bubbles is where non-members live.")}

{sec("01","Business Composition","Where does our membership come from?")}
<img class="img" src="data:image/png;base64,{charts['pie_county']}">
<img class="img" src="data:image/png;base64,{charts['pie_city']}">
{ins(f"Five counties account for {top5:.0f}% of all members. Buffalo and Rochester together represent ~22%. The bottom 10 counties contribute only {county_df.tail(10)['share'].sum():.1f}%. Growth must come from mid-tier deepening and dark-county activation.")}

{sec("02","Penetration — Best and Worst Markets","Which areas show the highest and lowest member penetration?")}
<div style="display:flex;gap:16px;flex-wrap:wrap;">
<div style="flex:1;min-width:45%;">
<h2 style="color:{GREEN};font-size:13px;">Best-Performing Cities (5K+ adults)</h2>
<table><tr><th>City</th><th>County</th><th style="text-align:right;">Pen %</th><th style="text-align:right;">Members</th><th style="text-align:right;">Adults</th></tr>"""

    for _,r in city_df[city_df["adults"]>5000].nlargest(12,"pen").iterrows():
        html+=f"<tr><td><strong>{r['city']}</strong></td><td>{r['county']}</td><td style='text-align:right;color:{GREEN};font-weight:700;'>{r['pen']}%</td><td style='text-align:right;'>{r['members']:,}</td><td style='text-align:right;'>{r['adults']:,}</td></tr>"

    html+=f"""</table></div>
<div style="flex:1;min-width:45%;">
<h2 style="color:{RED};font-size:13px;">Worst-Performing Cities (5K+ adults)</h2>
<table><tr><th>City</th><th>County</th><th style="text-align:right;">Pen %</th><th style="text-align:right;">Members</th><th style="text-align:right;">Adults</th></tr>"""

    for _,r in city_df[city_df["adults"]>5000].nsmallest(12,"pen").iterrows():
        html+=f"<tr><td><strong>{r['city']}</strong></td><td>{r['county']}</td><td style='text-align:right;color:{RED};font-weight:700;'>{r['pen']}%</td><td style='text-align:right;'>{r['members']:,}</td><td style='text-align:right;'>{r['adults']:,}</td></tr>"

    html+=f"""</table></div></div>
{ins("Winners share three traits: suburban location, income above $80K, and homeownership-heavy demographics. The worst performers are territory-edge cities with near-zero AAA visibility. Critically, low penetration does not follow low income — Binghamton ($55K income) has 0.2% penetration. This is a distribution gap, not a demand gap.")}

{sec("03","Top Opportunities — Where to Find New Members","Which cities have the largest untapped adult population?")}
<img class="img" src="data:image/png;base64,{charts['opp_cities']}">
{ins("The real action items are cities 4-12: Binghamton ($6M), Utica ($5.3M), Lockport ($3.8M), Ithaca ($3.7M), Endicott ($3.4M), Elmira ($3.2M), Jamestown ($3.0M). Large populations with near-zero penetration. Moving them from <1% to even 10% adds 40,000+ members.")}

{sec("04","Hidden Intelligence","What does the data reveal that isn't obvious?")}
<img class="img" src="data:image/png;base64,{charts['overunder']}">
{ins("<strong>Outperformers tell us what works:</strong> Churchville (+20.7pts), Lewiston (+19.4), Springville (+17.4) — mid-income markets exceeding what demographics predict. Something about local execution, agent presence, or community engagement is driving outsized results. Study and replicate.<br><br><strong>Underperformers reveal broken markets:</strong> Marcy (−33pts, $99K income), New Hartford (−31pts, $94K), Apalachin (−29pts, $86K) — wealthy communities with virtually no members. These households are not price-sensitive. They simply don't know AAA is there.")}

<img class="img" src="data:image/png;base64,{charts['donuts']}">
{ins("Inside our strongest counties, urban cores dramatically lag. Inner Buffalo (14215 = 9.9%) vs suburban (42-48%). Inner Rochester (14621 = 7.7%) vs Pittsford (47.9%). Different demographics require different product positioning — the Essentials tier and rental-community partnerships are designed for these neighborhoods.")}

{sec("05","Household Eligibility","How many households are eligible but not members?")}
<img class="img" src="data:image/png;base64,{charts['hh']}">
{ins(f"{hh:,} households in the territory. {T:,} have AAA membership — a {T/hh*100:.1f}% coverage rate. Even in Ontario (our best county), over half of households lack AAA. In Herkimer, 97% have no coverage. The ceiling is nowhere close.")}

{sec("06","Vehicle & Fleet Intelligence","How does vehicle penetration vary? What about aging fleets?")}
<img class="img" src="data:image/png;base64,{charts['veh']}">
<img class="img" src="data:image/png;base64,{charts['veh_age']}">
{ins(f"{veh:,} registered vehicles — {veh//T:.0f} per member. Even in our strongest counties, we cover only 30-38% of vehicles. {ov/veh*100:.0f}% of all vehicles are 3+ years old — the segment most needing roadside assistance and battery replacement.")}

<div style="margin:30px 0 12px;">
<div style="color:{NAVY};font-size:14px;font-weight:700;margin-bottom:4px;">Geographic View — Old vs New Fleet by County</div>
<div style="color:{DGRAY};font-size:11px;font-style:italic;">Left: red = highest aging fleet risk (most roadside opportunity). Right: green = counties with newest vehicles (EV/tech-forward demographics).</div>
</div>
<img class="img" src="data:image/png;base64,{charts['vehicle_geo_map']}">
{ins("Counties in the deep red (left map) carry 85–90% aging fleets and near-zero AAA penetration — the highest-priority battery and roadside acquisition zones. Counties in deep green (right map) skew toward newer, higher-value vehicles whose owners have greater interest in EV/roadside coverage products.")}

{sec("07","Battery & Roadside Opportunity","Where do aging vehicles meet low membership?")}
<img class="img" src="data:image/png;base64,{charts['battery']}">
{ins("409,985 old vehicles in markets with <10% penetration. At 10% battery service rate ($150 avg) = <strong>$6.1M service revenue</strong>. At 5% membership conversion = <strong>20,500 new members</strong>. Top targets: Lockport (35K old vehicles), Utica (35K), Elmira (26K), Jamestown (22K).")}

{sec("08","What Drives Membership?","Are there demographic patterns we can act on?")}
<img class="img" src="data:image/png;base64,{charts['income']}">
<img class="img" src="data:image/png;base64,{charts['demo']}">
{ins("Income (r=0.55+) is the strongest predictor. Over-$95K markets penetrate at 32.7% vs under-$50K at 11.3%. Education (r=0.40+) is second — Ithaca proves educated young people join at $65K income. A 'Finger Lakes Core' of high penetration radiates outward, decaying at every territorial edge regardless of income — a gravity model driven by awareness and presence.")}

{sec("09","County Performance at a Glance","Which counties are healthy vs broken — across every dimension?")}
<img class="img" src="data:image/png;base64,{charts['heatmap']}">
{ins("Read this left to right: green cells indicate strength, red indicates weakness. Counties at the top are leaders across all metrics. Counties near the bottom combine low penetration, low vehicle coverage, and aging fleets — the clearest candidates for market-entry investment. Note that high old-fleet percentage is shown as <strong>red</strong> (weakness = opportunity) because those vehicles need roadside service most.")}

{sec("10","Strategic Priority Matrix","Where should we invest — counties AND cities?")}
<img class="img" src="data:image/png;base64,{charts['quadrant_county']}">
<img class="img" src="data:image/png;base64,{charts['quadrant_city']}">
{ins("<strong>ENTER:</strong> Herkimer, Tioga, Chemung, Oneida + Binghamton, Utica, Elmira, Lockport — massive gaps, near-zero presence.<br><strong>GROW:</strong> Erie, Monroe, Onondaga, Niagara + Buffalo, Rochester, Syracuse — core markets, still 60-75% unconverted.<br><strong>PROTECT:</strong> Ontario, Madison, Genesee + Pittsford, East Amherst, Fairport — strong penetration, focus on retention.<br><strong>MONITOR:</strong> Smaller markets — maintain but don't over-invest.")}

{sec("11","Market Spotlight: Ithaca / Tompkins County","Targeted recommendations for a specific market")}
{ins("<strong>20,069 members | 80,486 adults | 24.7% penetration</strong><br>Income: $64,911 | Median age: 28.4 | College: 53%<br><br>Despite below-average income, Ithaca achieves 24.7% penetration driven by the highest education rate in the territory. Cornell + Ithaca College = 25,000+ students annually — a renewable acquisition pipeline. Recommendation: $39 student rate, campus ambassadors, digital-first enrollment, auto-convert at graduation.")}

{sec("12","Strategic Recommendations","Five initiatives for the next 12-24 months")}

{rec(1,"Activate Dark Counties","Herkimer, Tioga, Chemung, Oneida, Orleans, Chautauqua — 750K+ adults, <5% penetration",
    ["Satellite enrollment centers or local insurance agent partnerships",
     "'AAA Is Here' campaign: local radio, community events, direct mail to every household",
     "Introductory $49 first-year rate to break the awareness barrier",
     "Employer partnerships: Lockheed Martin (Owego), BAE Systems (Endicott), SUNY campuses"],
    "50,000 new members | $5.0M annual revenue", RED)}

{rec(2,"Aging Fleet Acquisition Campaign","409K old-vehicle owners in low-penetration markets",
    ["DMV-targeted direct mail with breakdown statistics",
     "Free battery check events at auto parts stores and tire shops",
     "Winter-readiness seasonal push (October-November)",
     "Used-car dealership bundling: 1-year AAA with every purchase"],
    "20,000 new members + $6.1M battery service revenue = $8.1M combined", ORANGE)}

{rec(3,"Close the Urban Donut Holes","Low-penetration zips within Erie, Monroe, Onondaga, Niagara",
    ["Neighborhood saturation in inner Buffalo (14215=9.9%), inner Rochester (14621=7.7%)",
     "Member referral bonus: $20 per successful referral",
     "Apartment complex and property manager partnerships",
     "Multi-product bundle: roadside + insurance + travel at 15% discount"],
    "Move top 4 counties from ~28% to 33% = 95,000 new members | $9.5M revenue", GREEN)}

{rec(4,"Student & Young Professional Pipeline","College towns: Ithaca, Syracuse, Buffalo, Rochester",
    ["$39/year student membership with app-based instant enrollment",
     "Campus ambassador program: 10 paid reps per school",
     "Auto-convert to full membership at graduation with 30-day free trial",
     "EV-ready messaging for environmentally conscious young adults"],
    "5,000 student members/year, 40% converting post-graduation", BLUE)}

{rec(5,"AAA Essentials + Battery Service","Lower-income markets + territory-wide service line",
    ["$39/year roadside-only tier for price-sensitive households",
     "$5/month no-commitment subscription option",
     "Mobile battery testing + replacement as paid non-member service ($150 avg)",
     "Community group rates through churches, civic organizations, county fairs"],
    "25,000 Essentials members ($975K) + $6M battery service = $7M combined", PURPLE)}

<div style="margin:35px 0;padding:25px;background:{LGRAY};border-radius:6px;text-align:center;border-top:3px solid {NAVY};">
<div style="color:{NAVY};font-size:18px;font-weight:800;">Combined Impact: +210,000 members &rarr; 1,053,000 total</div>
<div style="color:{DGRAY};font-size:13px;margin-top:4px;">+$29.6M incremental annual revenue</div>
</div>

<div style="page-break-before:always;margin-top:20px;padding:30px 30px 28px;background:{LGRAY};border:1px solid #E0E0E0;border-radius:6px;">
<div style="border-bottom:2px solid {NAVY};padding-bottom:8px;margin-bottom:16px;">
<span style="color:{TEAL};font-size:11px;font-weight:700;letter-spacing:1px;">APPENDIX</span>
<div style="color:{NAVY};font-size:18px;font-weight:700;margin-top:4px;">Data Sources &amp; Methodology</div>
</div>
<p style="font-size:11px;color:{DARK};line-height:1.75;margin:0 0 10px;">
<strong style="color:{NAVY};">Data Sources —</strong> AAA WCNY membership records, U.S. Census Bureau demographics, and New York
State DMV vehicle registrations. Penetration = members &divide; adults 18+. Revenue opportunity sized at
$100 average annual membership value. All figures as of April 2026.
</p>
<p style="font-size:11px;color:{DARK};line-height:1.75;margin:0 0 10px;">
<strong style="color:{NAVY};">Membership Data —</strong> Member counts by ZIP code sourced from AAA WCNY internal records.
County and city roll-ups aggregated from ZIP-level data. Total member count held at 843,000.
</p>
<p style="font-size:11px;color:{DARK};line-height:1.75;margin:0 0 10px;">
<strong style="color:{NAVY};">Demographic Data —</strong> U.S. Census Bureau American Community Survey (ACS) 5-Year Estimates.
Adults 18+ used as the penetration denominator (not total population).
Household income, median age, college education rates from ACS ZIP-level tables.
</p>
<p style="font-size:11px;color:{DARK};line-height:1.75;margin:0 0 10px;">
<strong style="color:{NAVY};">Vehicle Data —</strong> New York State Department of Motor Vehicles vehicle registration data by ZIP code.
Old vehicles defined as registered 3+ years prior to data date. New vehicles registered within 3 years.
</p>
<p style="font-size:11px;color:{DARK};line-height:1.75;margin:0;">
<strong style="color:{NAVY};">Study Intent —</strong> This analysis answers ten strategic questions about membership penetration:
where we are strongest and weakest, what drives the differences, and what specific actions can move
AAA WCNY from 843,000 members toward one million. Designed for board-level resource allocation
and go-to-market decisions over the next 12–24 months.
</p>
</div>

<div style="text-align:center;color:{MGRAY};font-size:8px;padding:30px 0;border-top:1px solid #E0E0E0;margin-top:40px;">
AAA WCNY &bull; Membership Penetration Analysis &bull; Confidential &bull; {today}
</div>
</div></body></html>"""
    return html


def main():
    print("Loading data...")
    mem,census,dmv = load()
    zdf = build_zip_df(mem,census,dmv)
    city_df = agg(zdf, ["city","county"])
    county_df = agg(zdf, ["county"]).sort_values("members", ascending=False)

    print("  Loading geo data...")
    county_geo, zip_latlng = load_geo()

    charts={}
    steps=[
        ("county_map", lambda: chart_county_choropleth(county_df, county_geo)),
        ("city_map", lambda: chart_city_bubble_map(city_df, county_geo, zip_latlng)),
        ("pie_county", lambda: chart_pie_county(county_df)),
        ("pie_city", lambda: chart_pie_city(city_df)),
        ("opp_cities", lambda: chart_opp_cities(city_df)),
        ("veh", lambda: chart_veh_vs_adult(county_df)),
        ("vehicle_geo_map", lambda: chart_vehicle_geo_map(county_df, county_geo)),
        ("battery", lambda: chart_battery(city_df)),
        ("income", lambda: chart_income(zdf)),
        ("quadrant_county", lambda: chart_quadrant_county(county_df)),
        ("quadrant_city", lambda: chart_city_strategy(city_df)),
        ("heatmap", lambda: chart_heatmap(county_df)),
        ("demo", lambda: chart_demo(zdf)),
        ("hh", lambda: chart_hh(county_df)),
        ("veh_age", lambda: chart_veh_age(county_df)),
        ("overunder", lambda: chart_overunder(zdf)),
        ("donuts", lambda: chart_donuts(zdf)),
    ]
    for name,fn in steps:
        print(f"  {name}...")
        charts[name]=fn()

    print("Building HTML...")
    html=build(zdf, city_df, county_df, charts)
    out=os.path.join(ROOT,"AAA_WCNY_Membership_Analysis.html")
    with open(out,"w") as f: f.write(html)
    print(f"Done: {out}")

if __name__=="__main__": main()
