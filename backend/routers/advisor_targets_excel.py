"""Advisor Targets Excel import and export endpoints mimicking the All Advisors template."""

import io
import logging
from datetime import datetime
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from database import get_db
from models import AdvisorTarget, MonthlyAdvisorTarget, TargetUpload, User
from auth import get_current_user, require_admin
from activity_logger import log_activity
from routers.advisor_targets_monthly import get_monthly_targets
from routers.advisor_targets_helpers import _get_comm_rate_accurate
from routers.advisor_targets import _normalize_name
from shared import WON_STAGES, line_filter_opp

router = APIRouter()
log = logging.getLogger('salesinsight.targets.excel')


def _parse_target_value(val) -> float:
    """Parse numeric cell values from spreadsheet."""
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace('$', '').replace(',', '')
    if not s or s.lower() in ('none', 'no targets', '0', '-'):
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


@router.get("/api/targets/monthly/{year}/export")
def export_monthly_targets_excel(
    year: int,
    line: str = "Travel",
    base: str = "commission",  # commission or bookings
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Export monthly targets matching the 6-row-per-advisor 'All Advisors' layout with formulas."""
    # 1. Fetch structured advisor targets and actuals
    res = get_monthly_targets(year=year, line=line, _user=user, db=db)
    advisors = res.get('advisors', [])

    # 2. Create openpyxl workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "All Advisors"
    ws.views.sheetView[0].showGridLines = True

    # 3. Define styling
    font_family = "Segoe UI"
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")  # Charcoal
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    
    label_font = Font(name=font_family, size=10, bold=False, color="000000")
    bold_label_font = Font(name=font_family, size=10, bold=True, color="000000")
    italic_label_font = Font(name=font_family, size=10, italic=True, color="555555")

    thin_border = Border(
        left=Side(style='thin', color='E5E7EB'),
        right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'),
        bottom=Side(style='thin', color='E5E7EB')
    )
    
    double_bottom_border = Border(
        left=Side(style='thin', color='E5E7EB'),
        right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'),
        bottom=Side(style='double', color='000000')  # Thick group break
    )

    actuals_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")  # Off-white
    pos_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")      # Soft Green
    neg_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")      # Soft Red

    # Number formats
    fmt_currency = "$#,##0"
    fmt_percentage = "0.0%"

    # 4. Headers definition
    headers = [
        "Associate", "Title", "Branch", "Annual Performance Threshold", 
        "Monthly Performance Threshold", "Stretch to Branch Budget Target", 
        "January", "February", "March", "Q1 Stretch Goal", "", 
        "April", "May", "June", "Q2 Stretch Goal", 
        "July", "August", "September", "Q3 Stretch Goal", 
        "October", "November", "December", "Q4 Stretch Goal", 
        "Sum of Qtrly Stretch", "Year End Stretch Goal"
    ]

    # Write headers
    for col_idx, text in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center" if col_idx > 6 else "left", vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 28

    # 5. Populate advisor blocks
    current_row = 2

    for adv in advisors:
        # Determine values depending on target base (commission or bookings)
        # Months are 1-12. Target amounts are retrieved from 'months' array
        is_comm = (base == 'commission')
        
        # Pull actual values
        months_cy = [
            m['bookings_actual'] if not is_comm else m['actual']
            for m in adv['months']
        ]
        months_py = [
            m['bookings_actual_py'] if not is_comm else m['actual_py']
            for m in adv['months']
        ]
        
        # Targets are in adv['months']
        months_tgt = [
            m['target_bookings'] if not is_comm else m['target']
            for m in adv['months']
        ]

        # Define 6 rows for this advisor block
        R = current_row

        # --- Row R: Target (Stretch Goal) ---
        ws.cell(row=R, column=1, value=adv['name']).font = bold_label_font
        ws.cell(row=R, column=2, value=adv['title']).font = label_font
        ws.cell(row=R, column=3, value=adv['branch']).font = label_font
        
        # Annual/Monthly Thresholds
        ann_thresh = adv['annual_threshold']
        if not is_comm:
            # Bookings threshold is Annual Commission Threshold / avg_comm_rate
            comm_rate = res['methodology']['commission_rate'] / 100.0
            ann_thresh = round(ann_thresh / comm_rate) if comm_rate > 0 else ann_thresh

        ws.cell(row=R, column=4, value=ann_thresh).font = label_font
        ws.cell(row=R, column=4).number_format = fmt_currency
        
        # Monthly Performance Threshold formula
        ws.cell(row=R, column=5, value=f"=D{R}/12").font = label_font
        ws.cell(row=R, column=5).number_format = fmt_currency

        # Stretch Target
        stretch_total = sum(months_tgt)
        ws.cell(row=R, column=6, value=stretch_total).font = bold_label_font
        ws.cell(row=R, column=6).number_format = fmt_currency

        # Month Columns (Target values)
        for i, val in enumerate(months_tgt):
            col = [7, 8, 9, 12, 13, 14, 16, 17, 18, 20, 21, 22][i]
            c = ws.cell(row=R, column=col, value=val)
            c.font = label_font
            c.number_format = fmt_currency

        # Quarterly Sum Formulas
        ws.cell(row=R, column=10, value=f"=SUM(G{R}:I{R})").font = bold_label_font
        ws.cell(row=R, column=10).number_format = fmt_currency
        ws.cell(row=R, column=15, value=f"=SUM(L{R}:N{R})").font = bold_label_font
        ws.cell(row=R, column=15).number_format = fmt_currency
        ws.cell(row=R, column=19, value=f"=SUM(P{R}:R{R})").font = bold_label_font
        ws.cell(row=R, column=19).number_format = fmt_currency
        ws.cell(row=R, column=23, value=f"=SUM(T{R}:V{R})").font = bold_label_font
        ws.cell(row=R, column=23).number_format = fmt_currency

        # Column X: Sum of Qtrly Stretch
        ws.cell(row=R, column=24, value=f"=SUM(J{R},O{R},S{R},W{R})").font = bold_label_font
        ws.cell(row=R, column=24).number_format = fmt_currency
        
        # Column Y: Year End Stretch Goal
        ws.cell(row=R, column=25, value=f"=F{R}").font = bold_label_font
        ws.cell(row=R, column=25).number_format = fmt_currency

        # --- Row R+1: ACTUALS {year} ---
        ws.cell(row=R+1, column=5, value=f"ACTUALS {year}").font = bold_label_font
        
        # Actual values
        for i, val in enumerate(months_cy):
            col = [7, 8, 9, 12, 13, 14, 16, 17, 18, 20, 21, 22][i]
            c = ws.cell(row=R+1, column=col, value=val)
            c.font = label_font
            c.number_format = fmt_currency
            c.fill = actuals_fill

        # Actual Qtrly Sums
        ws.cell(row=R+1, column=10, value=f"=SUM(G{R+1}:I{R+1})").font = bold_label_font
        ws.cell(row=R+1, column=10).number_format = fmt_currency
        ws.cell(row=R+1, column=15, value=f"=SUM(L{R+1}:N{R+1})").font = bold_label_font
        ws.cell(row=R+1, column=15).number_format = fmt_currency
        ws.cell(row=R+1, column=19, value=f"=SUM(P{R+1}:R{R+1})").font = bold_label_font
        ws.cell(row=R+1, column=19).number_format = fmt_currency
        ws.cell(row=R+1, column=23, value=f"=SUM(T{R+1}:V{R+1})").font = bold_label_font
        ws.cell(row=R+1, column=23).number_format = fmt_currency
        
        # Year End Actual
        ws.cell(row=R+1, column=25, value=f"=SUM(J{R+1},O{R+1},S{R+1},W{R+1})").font = bold_label_font
        ws.cell(row=R+1, column=25).number_format = fmt_currency

        # --- Row R+2: Actuals {year-1} ---
        ws.cell(row=R+2, column=5, value=f"Actuals {year-1}").font = italic_label_font

        for i, val in enumerate(months_py):
            col = [7, 8, 9, 12, 13, 14, 16, 17, 18, 20, 21, 22][i]
            c = ws.cell(row=R+2, column=col, value=val)
            c.font = label_font
            c.number_format = fmt_currency

        # Qtrly Sums
        ws.cell(row=R+2, column=10, value=f"=SUM(G{R+2}:I{R+2})").font = label_font
        ws.cell(row=R+2, column=10).number_format = fmt_currency
        ws.cell(row=R+2, column=15, value=f"=SUM(L{R+2}:N{R+2})").font = label_font
        ws.cell(row=R+2, column=15).number_format = fmt_currency
        ws.cell(row=R+2, column=19, value=f"=SUM(P{R+2}:R{R+2})").font = label_font
        ws.cell(row=R+2, column=19).number_format = fmt_currency
        ws.cell(row=R+2, column=23, value=f"=SUM(T{R+2}:V{R+2})").font = label_font
        ws.cell(row=R+2, column=23).number_format = fmt_currency
        
        # Year End Actual
        ws.cell(row=R+2, column=25, value=f"=SUM(J{R+2},O{R+2},S{R+2},W{R+2})").font = label_font
        ws.cell(row=R+2, column=25).number_format = fmt_currency

        # --- Row R+3: Variance YOY ---
        ws.cell(row=R+3, column=5, value="Variance YOY").font = label_font

        for col in [7, 8, 9, 12, 13, 14, 16, 17, 18, 20, 21, 22]:
            col_letter = get_column_letter(col)
            c = ws.cell(row=R+3, column=col, value=f"={col_letter}{R+1}-{col_letter}{R+2}")
            c.font = label_font
            c.number_format = fmt_currency
            
            # Highlight variance
            act_val = months_cy[[7, 8, 9, 12, 13, 14, 16, 17, 18, 20, 21, 22].index(col)]
            py_val = months_py[[7, 8, 9, 12, 13, 14, 16, 17, 18, 20, 21, 22].index(col)]
            c.fill = pos_fill if act_val >= py_val else neg_fill

        # Qtrly Sum Variances
        for q_col, q_tgt_col in [(10, "J"), (15, "O"), (19, "S"), (23, "W")]:
            c = ws.cell(row=R+3, column=q_col, value=f"={q_tgt_col}{R+1}-{q_tgt_col}{R+2}")
            c.font = bold_label_font
            c.number_format = fmt_currency

        # Q1 YoY% in Col 11
        ws.cell(row=R+3, column=11, value=f"=IF(J{R+2}=0,0,(J{R+1}-J{R+2})/J{R+2})").font = bold_label_font
        ws.cell(row=R+3, column=11).number_format = fmt_percentage

        # Year End YoY Variance
        ws.cell(row=R+3, column=25, value=f"=Y{R+1}-Y{R+2}").font = bold_label_font
        ws.cell(row=R+3, column=25).number_format = fmt_currency

        # --- Row R+4: Variance to Perf. Thresh ---
        ws.cell(row=R+4, column=5, value="Variance to Perf. Thresh").font = label_font

        for col in [7, 8, 9, 12, 13, 14, 16, 17, 18, 20, 21, 22]:
            col_letter = get_column_letter(col)
            c = ws.cell(row=R+4, column=col, value=f"=$E{R}-{col_letter}{R+1}")
            c.font = label_font
            c.number_format = fmt_currency
            
            # Highlight variance (Note: positive variance here means actual is LESS than threshold, which might be negative context.
            # But usually we highlight soft-red if actuals are LESS than threshold. So if E14-G15 > 0, it means actuals are BELOW threshold. Let's color accordingly.)
            act_val = months_cy[[7, 8, 9, 12, 13, 14, 16, 17, 18, 20, 21, 22].index(col)]
            mo_threshold = ann_thresh / 12
            c.fill = pos_fill if act_val >= mo_threshold else neg_fill

        # Qtrly Threshold Variances
        ws.cell(row=R+4, column=10, value=f"=($E{R}*3)-J{R+1}").font = bold_label_font
        ws.cell(row=R+4, column=10).number_format = fmt_currency
        ws.cell(row=R+4, column=15, value=f"=($E{R}*3)-O{R+1}").font = bold_label_font
        ws.cell(row=R+4, column=15).number_format = fmt_currency
        ws.cell(row=R+4, column=19, value=f"=($E{R}*3)-S{R+1}").font = bold_label_font
        ws.cell(row=R+4, column=19).number_format = fmt_currency
        ws.cell(row=R+4, column=23, value=f"=($E{R}*3)-W{R+1}").font = bold_label_font
        ws.cell(row=R+4, column=23).number_format = fmt_currency

        # Sum of Qtrly Threshold Variances (Col X)
        ws.cell(row=R+4, column=24, value=f"=SUM(J{R+4},O{R+4},S{R+4},W{R+4})").font = bold_label_font
        ws.cell(row=R+4, column=24).number_format = fmt_currency

        # Year End Threshold Variance (Col Y)
        ws.cell(row=R+4, column=25, value=f"=D{R}-Y{R+1}").font = bold_label_font
        ws.cell(row=R+4, column=25).number_format = fmt_currency

        # --- Row R+5: Variance to Stretch ---
        ws.cell(row=R+5, column=5, value="Variance to Stretch").font = bold_label_font

        for col in [7, 8, 9, 12, 13, 14, 16, 17, 18, 20, 21, 22]:
            col_letter = get_column_letter(col)
            c = ws.cell(row=R+5, column=col, value=f"={col_letter}{R}-{col_letter}{R+1}")
            c.font = label_font
            c.number_format = fmt_currency
            
            act_val = months_cy[[7, 8, 9, 12, 13, 14, 16, 17, 18, 20, 21, 22].index(col)]
            tgt_val = months_tgt[[7, 8, 9, 12, 13, 14, 16, 17, 18, 20, 21, 22].index(col)]
            c.fill = pos_fill if act_val >= tgt_val else neg_fill

        # Qtrly Sum Variances
        ws.cell(row=R+5, column=10, value=f"=J{R}-J{R+1}").font = bold_label_font
        ws.cell(row=R+5, column=10).number_format = fmt_currency
        ws.cell(row=R+5, column=15, value=f"=O{R}-O{R+1}").font = bold_label_font
        ws.cell(row=R+5, column=15).number_format = fmt_currency
        ws.cell(row=R+5, column=19, value=f"=S{R}-S{R+1}").font = bold_label_font
        ws.cell(row=R+5, column=19).number_format = fmt_currency
        ws.cell(row=R+5, column=23, value=f"=W{R}-W{R+1}").font = bold_label_font
        ws.cell(row=R+5, column=23).number_format = fmt_currency

        # Sum of Qtrly Stretch Variances (Col X)
        ws.cell(row=R+5, column=24, value=f"=X{R}-X{R+1}").font = bold_label_font
        ws.cell(row=R+5, column=24).number_format = fmt_currency

        # Year End Stretch Variance (Col Y)
        ws.cell(row=R+5, column=25, value=f"=Y{R}-Y{R+1}").font = bold_label_font
        ws.cell(row=R+5, column=25).number_format = fmt_currency

        # --- Set border structures ---
        for r_offset in range(6):
            r_idx = R + r_offset
            ws.row_dimensions[r_idx].height = 19
            for c_idx in range(1, 26):
                cell = ws.cell(row=r_idx, column=c_idx)
                # Apply thin borders inside the block, double border on the bottom of the 6th row
                cell.border = double_bottom_border if r_offset == 5 else thin_border
                # Align numbers center/right
                if c_idx > 6:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                elif c_idx in (4, 5, 6):
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

        current_row += 6

    # 6. Apply Column Autofits
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            # Skip formula cells for length check to prevent long formula strings blowing up column width
            val = str(cell.value or '')
            if val.startswith('='):
                continue
            max_len = max(max_len, len(val))
        
        # Set generous widths based on column contents
        ws.column_dimensions[col_letter].width = max(max_len + 3, 13)

    # Specific override for names/titles/branches
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 24
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 24
    ws.column_dimensions['F'].width = 18

    # 7. Write to memory buffer and stream back
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)

    filename = f"Monthly_Targets_{line}_{year}_{base}.xlsx"
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.post("/api/targets/monthly/{year}/import")
async def import_monthly_targets_excel(
    year: int,
    line: str = "Travel",
    base: str = "commission",
    file: UploadFile = File(...),
    admin: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Import monthly targets from spreadsheet (supports 6-row block template and flat 1-row format)."""
    if not file.filename:
        raise HTTPException(status_code=400, detail="No file uploaded")

    content = await file.read()
    if len(content) > 5 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File too large (max 5MB)")

    try:
        # Load workbook with data_only=True to evaluate all formulas!
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception as e:
        log.error(f"Failed to parse uploaded Excel: {e}")
        raise HTTPException(status_code=400, detail="Invalid Excel workbook format")

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    if len(rows) < 2:
        raise HTTPException(status_code=400, detail="Spreadsheet has no data rows")

    # Fetch commission rate
    lf = line_filter_opp(line)
    from sf_client import sf_query_all
    import cache
    comm_rate = _get_comm_rate_accurate(line, year - 1, cache, sf_query_all, WON_STAGES, lf)
    if comm_rate <= 0:
        comm_rate = 0.187  # default fallback

    # Get system upload ID
    upload = db.query(TargetUpload).filter(TargetUpload.filename == '__sf_auto__').first()
    upload_id = upload.id if upload else 1

    # Load all existing advisor targets to perform fuzzy name matching in Python
    all_advisors = db.query(AdvisorTarget).all()
    advisor_map = {}
    for adv in all_advisors:
        if adv.sf_name:
            advisor_map[_normalize_name(adv.sf_name).lower()] = adv
        if adv.raw_name:
            advisor_map[_normalize_name(adv.raw_name).lower()] = adv

    advisor_updates = 0
    targets_updated = 0

    # Auto-detect format: check if column E (index 4) has block labels like ACTUALS/Variance/Stretch
    is_six_row_format = False
    if len(rows) >= 7:
        for r_idx in range(2, min(7, len(rows))):
            row_cells = rows[r_idx]
            if len(row_cells) > 4:
                val = str(row_cells[4] or '').strip().upper()
                if 'ACTUAL' in val or 'VAR' in val or 'STRETCH' in val:
                    is_six_row_format = True
                    break

    if is_six_row_format:
        # Format A: 6-row block per advisor
        month_columns = [6, 7, 8, 11, 12, 13, 15, 16, 17, 19, 20, 21]
        R = 1
        total_rows = len(rows)
        
        while R < total_rows:
            row_data = rows[R]
            raw_name = str(row_data[0] or '').strip()
            if not raw_name:
                R += 6
                continue

            title_val = str(row_data[1] or '').strip() or None
            branch_val = str(row_data[2] or '').strip() or None
            
            annual_threshold_val = _parse_target_value(row_data[3])
            annual_stretch_val = _parse_target_value(row_data[5])

            sf_name = _normalize_name(raw_name)
            advisor = advisor_map.get(sf_name.lower()) or advisor_map.get(raw_name.lower())

            if not advisor:
                advisor = AdvisorTarget(
                    upload_id=upload_id,
                    raw_name=raw_name,
                    sf_name=sf_name,
                    branch=branch_val,
                    title=title_val,
                    monthly_target=round(annual_threshold_val / 12) if annual_threshold_val > 0 else None,
                    annual_stretch=annual_stretch_val if annual_stretch_val > 0 else None,
                )
                db.add(advisor)
                db.flush()
                # Update map to prevent future duplicates
                advisor_map[sf_name.lower()] = advisor
                advisor_map[raw_name.lower()] = advisor
            else:
                advisor.title = title_val or advisor.title
                advisor.branch = branch_val or advisor.branch
                if annual_threshold_val > 0:
                    advisor.monthly_target = round(annual_threshold_val / 12)
                if annual_stretch_val > 0:
                    advisor.annual_stretch = annual_stretch_val
                db.add(advisor)

            advisor_updates += 1

            for month_idx, col_idx in enumerate(month_columns, 1):
                if col_idx < len(row_data):
                    val = _parse_target_value(row_data[col_idx])
                    
                    if base == 'bookings':
                        bookings_val = round(val)
                        commission_val = round(val * comm_rate)
                    else:
                        commission_val = round(val)
                        bookings_val = round(val / comm_rate) if comm_rate > 0 else round(val)

                    existing = db.query(MonthlyAdvisorTarget).filter(
                        MonthlyAdvisorTarget.advisor_target_id == advisor.id,
                        MonthlyAdvisorTarget.year == year,
                        MonthlyAdvisorTarget.month == month_idx,
                    ).first()

                    if existing:
                        existing.target_amount = commission_val
                        existing.target_bookings = bookings_val
                        existing.updated_by_email = admin.email
                        existing.updated_at = datetime.utcnow()
                    else:
                        db.add(MonthlyAdvisorTarget(
                            advisor_target_id=advisor.id,
                            year=year,
                            month=month_idx,
                            target_amount=commission_val,
                            target_bookings=bookings_val,
                            updated_by_email=admin.email,
                        ))
                    targets_updated += 1

            R += 6
    else:
        # Format B: Flat 1-row per advisor with January to December columns
        from routers.advisor_targets import MONTH_PATTERNS
        headers = [str(h or '').strip() for h in rows[0]]
        mapping = {}
        
        for idx, h in enumerate(headers):
            hl = h.lower().strip()
            if not mapping.get('name') and any(k in hl for k in ('name', 'employee', 'advisor', 'agent', 'associate')):
                mapping['name'] = idx
            elif not mapping.get('branch') and any(k in hl for k in ('branch', 'office', 'location')):
                mapping['branch'] = idx
            elif not mapping.get('title') and any(k in hl for k in ('title', 'position', 'role')):
                mapping['title'] = idx
            elif not mapping.get('annual_threshold') and any(k in hl for k in ('annual threshold', 'performance threshold', 'annual performance')):
                mapping['annual_threshold'] = idx
            elif not mapping.get('annual_stretch') and any(k in hl for k in ('stretch', 'budget', 'annual stretch')):
                mapping['annual_stretch'] = idx
            else:
                for pattern, month_num in MONTH_PATTERNS.items():
                    if hl.startswith(pattern):
                        key = f"m{month_num}"
                        if key not in mapping:
                            mapping[key] = idx
                        break

        if 'name' not in mapping:
            raise HTTPException(
                status_code=400,
                detail=f"Could not identify name/associate column. Found columns: {headers}",
            )

        R = 1
        total_rows = len(rows)
        while R < total_rows:
            row_data = rows[R]
            if not any(row_data):
                R += 1
                continue

            raw_name = str(row_data[mapping['name']] or '').strip() if mapping['name'] < len(row_data) else ''
            if not raw_name:
                R += 1
                continue

            branch_val = str(row_data[mapping['branch']] or '').strip() if mapping.get('branch') is not None and mapping['branch'] < len(row_data) else None
            title_val = str(row_data[mapping['title']] or '').strip() if mapping.get('title') is not None and mapping['title'] < len(row_data) else None
            
            annual_threshold_val = _parse_target_value(row_data[mapping['annual_threshold']]) if mapping.get('annual_threshold') is not None and mapping['annual_threshold'] < len(row_data) else 0.0
            annual_stretch_val = _parse_target_value(row_data[mapping['annual_stretch']]) if mapping.get('annual_stretch') is not None and mapping['annual_stretch'] < len(row_data) else 0.0

            sf_name = _normalize_name(raw_name)
            advisor = advisor_map.get(sf_name.lower()) or advisor_map.get(raw_name.lower())

            monthly_values = []
            for m in range(1, 13):
                col_idx = mapping.get(f"m{m}")
                if col_idx is not None and col_idx < len(row_data):
                    val = _parse_target_value(row_data[col_idx])
                    monthly_values.append(val)
                else:
                    monthly_values.append(0.0)

            if annual_stretch_val == 0.0 and sum(monthly_values) > 0:
                annual_stretch_val = sum(monthly_values)

            if not advisor:
                advisor = AdvisorTarget(
                    upload_id=upload_id,
                    raw_name=raw_name,
                    sf_name=sf_name,
                    branch=branch_val,
                    title=title_val,
                    monthly_target=round(annual_threshold_val / 12) if annual_threshold_val > 0 else None,
                    annual_stretch=annual_stretch_val if annual_stretch_val > 0 else None,
                )
                db.add(advisor)
                db.flush()
                advisor_map[sf_name.lower()] = advisor
                advisor_map[raw_name.lower()] = advisor
            else:
                advisor.title = title_val or advisor.title
                advisor.branch = branch_val or advisor.branch
                if annual_threshold_val > 0:
                    advisor.monthly_target = round(annual_threshold_val / 12)
                if annual_stretch_val > 0:
                    advisor.annual_stretch = annual_stretch_val
                db.add(advisor)

            advisor_updates += 1

            for month_idx in range(1, 13):
                val = monthly_values[month_idx - 1]
                
                if base == 'bookings':
                    bookings_val = round(val)
                    commission_val = round(val * comm_rate)
                else:
                    commission_val = round(val)
                    bookings_val = round(val / comm_rate) if comm_rate > 0 else round(val)

                existing = db.query(MonthlyAdvisorTarget).filter(
                    MonthlyAdvisorTarget.advisor_target_id == advisor.id,
                    MonthlyAdvisorTarget.year == year,
                    MonthlyAdvisorTarget.month == month_idx,
                ).first()

                if existing:
                    existing.target_amount = commission_val
                    existing.target_bookings = bookings_val
                    existing.updated_by_email = admin.email
                    existing.updated_at = datetime.utcnow()
                else:
                    db.add(MonthlyAdvisorTarget(
                        advisor_target_id=advisor.id,
                        year=year,
                        month=month_idx,
                        target_amount=commission_val,
                        target_bookings=bookings_val,
                        updated_by_email=admin.email,
                    ))
                targets_updated += 1

            R += 1

    db.commit()

    log_activity(
        db, action='targets_excel_imported', category='targets',
        user=admin,
        detail=f"Imported Excel targets file for {year} (advisors={advisor_updates}, base={base})",
        metadata={'year': year, 'advisors': advisor_updates, 'cells': targets_updated, 'base': base}
    )

    return {
        'status': 'success',
        'advisors_updated': advisor_updates,
        'targets_updated': targets_updated,
    }
