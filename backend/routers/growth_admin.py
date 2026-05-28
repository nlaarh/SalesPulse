"""
Growth Data Refresh — pulls fresh data from Databricks and updates seed CSVs.

Admin-only endpoint. Re-runs the same queries as Analysis_Apr/strategic_report/databricks_loader.py
but writes to backend/seed_data/growth_data/ so the growth dashboard serves live data.

Source tables:
  - dev_bronze_catalog.legacy_datawarehouse.membership_consolidated
  - dev_silver_catalog.data_warehouse.insurance_client_accounts_f
  - dev_silver_catalog.data_warehouse.insurance_policies_f
  - dev_silver_catalog.data_warehouse.travel_store_transactions_f
  - dev_silver_catalog.data_warehouse.battery_test_transactions_f
  - dev_silver_catalog.data_warehouse.ers_calls_details
  - dev_gold_catalog.business_intelligence.crm_account
"""
from __future__ import annotations
import json
import os
import shutil
import time
from datetime import date, datetime
from pathlib import Path

from fastapi import APIRouter, Depends, BackgroundTasks, UploadFile, File
from dotenv import load_dotenv

from auth import require_admin
from routers.growth_queries import (
    _territory_zips,
    ALL_DATASETS, DATASET_CSV_MAP,
)

router = APIRouter(tags=["growth-admin"])

SEED_DIR = Path(__file__).resolve().parent.parent / "seed_data"
DATA_DIR = SEED_DIR / "growth_data"
DATA_DIR.mkdir(parents=True, exist_ok=True)
BACKUP_DIR = SEED_DIR / "growth_backups"
BACKUP_DIR.mkdir(parents=True, exist_ok=True)

load_dotenv(Path(__file__).resolve().parents[2] / ".env")

_refresh_status = {"running": False, "last_run": None, "last_error": None, "results": {}}

# Datasets whose refresh function takes no zips argument (global territory-wide)
_NO_ZIP_DATASETS = {"membership_trend"}


# ─── Helpers ─────────────────────────────────────────────────────────────

def _backup_current_data():
    """Backup existing CSVs to timestamped folder before overwriting."""
    existing_csvs = list(DATA_DIR.glob("*.csv"))
    if not existing_csvs:
        return None
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = BACKUP_DIR / stamp
    backup_path.mkdir(parents=True, exist_ok=True)
    for f in existing_csvs:
        shutil.copy2(f, backup_path / f.name)
    for f in DATA_DIR.glob("*.json"):
        shutil.copy2(f, backup_path / f.name)
    all_backups = sorted(BACKUP_DIR.iterdir(), reverse=True)
    for old in all_backups[10:]:
        if old.is_dir():
            shutil.rmtree(old)
    return stamp


def _clear_disk_cache():
    """Clear all growth_ prefixed disk cache entries so next request regenerates from fresh CSVs."""
    import re as _re
    try:
        import cache as _cache
        import json as _json
        cleared = 0
        for entry in _cache._CACHE_DIR.glob('*.json'):
            try:
                payload = _json.loads(entry.read_text())
                if _re.match(r'^growth_', payload.get('key', '')):
                    entry.unlink()
                    cleared += 1
            except Exception:
                pass
        return cleared
    except Exception:
        return 0


def _do_refresh():
    """Run all Databricks queries and update CSVs. Called as background task."""
    global _refresh_status
    _refresh_status["running"] = True
    _refresh_status["results"] = {}
    _refresh_status["last_error"] = None

    backup_stamp = _backup_current_data()
    if backup_stamp:
        _refresh_status["results"]["_backup"] = {"folder": backup_stamp, "status": "ok"}

    zips = _territory_zips()
    if not zips:
        _refresh_status["running"] = False
        _refresh_status["last_error"] = "No territory ZIPs found in territory_zips.json"
        return

    for name, fn in ALL_DATASETS:
        try:
            start = time.time()
            count = fn() if name in _NO_ZIP_DATASETS else fn(zips)
            elapsed = round(time.time() - start, 1)
            _refresh_status["results"][name] = {"rows": count, "seconds": elapsed, "status": "ok"}
        except Exception as e:
            _refresh_status["results"][name] = {"rows": 0, "seconds": 0, "status": "error", "error": str(e)[:200]}

    _refresh_status["running"] = False
    _refresh_status["last_run"] = date.today().isoformat()

    # Clear LRU cache so growth router re-reads CSVs on next request
    from routers.growth import _build_zip_table
    _build_zip_table.cache_clear()

    # Clear disk cache so next HTTP request regenerates from fresh CSVs (not 30-day stale)
    _clear_disk_cache()


# ─── API Endpoints ───────────────────────────────────────────────────────

@router.get("/api/growth/data-status")
def data_status(_user=Depends(require_admin)):
    """Check current data freshness and refresh status."""
    files = {}
    for f in DATA_DIR.glob("*.csv"):
        stat = f.stat()
        files[f.stem] = {
            "file": f.name,
            "size_bytes": stat.st_size,
            "last_modified": time.strftime("%Y-%m-%d %H:%M", time.localtime(stat.st_mtime)),
        }

    has_databricks = all([
        os.environ.get("DATABRICKS_HOST"),
        os.environ.get("DATABRICKS_TOKEN"),
        os.environ.get("DATABRICKS_HTTP_PATH"),
    ])

    return {
        "files": files,
        "refresh_status": _refresh_status,
        "databricks_configured": has_databricks,
    }


@router.post("/api/growth/refresh")
def trigger_refresh(background_tasks: BackgroundTasks, _user=Depends(require_admin)):
    """Trigger a background data refresh from Databricks."""
    if _refresh_status["running"]:
        return {"status": "already_running", "message": "Refresh is already in progress"}

    has_databricks = all([
        os.environ.get("DATABRICKS_HOST"),
        os.environ.get("DATABRICKS_TOKEN"),
        os.environ.get("DATABRICKS_HTTP_PATH"),
    ])
    if not has_databricks:
        return {"status": "error", "message": "Databricks credentials not configured (DATABRICKS_HOST, DATABRICKS_TOKEN, DATABRICKS_HTTP_PATH)"}

    background_tasks.add_task(_do_refresh)
    return {"status": "started", "message": "Data refresh started. Check /api/growth/data-status for progress."}


@router.post("/api/growth/upload-census")
async def upload_census(file: UploadFile = File(...), _user=Depends(require_admin)):
    """
    Upload a new Census_Data_CustomerSeg_Vehicles.xlsx to refresh demographic data.

    This data comes from US Census Bureau (population, age, housing, income)
    and NY DMV (registered vehicles). Updated annually.
    """
    import openpyxl
    from io import BytesIO

    content = await file.read()
    wb = openpyxl.load_workbook(BytesIO(content), data_only=True, read_only=True)

    sheet_name = None
    for name in wb.sheetnames:
        if "census" in name.lower() or "custseg" in name.lower():
            sheet_name = name
            break
    if not sheet_name:
        sheet_name = wb.sheetnames[0]

    ws = wb[sheet_name]
    keys = [
        "zip_code", "city", "county", "coverage", "region",
        "registered_vehicles", "vehicles_3plus_yrs",
        "owner_occupied", "untapped_homes", "renter_occupied",
        "population", "adults_18plus", "median_income", "median_home_value",
        "age_16_18", "age_18_24", "age_25_34", "age_35_44", "age_45_54",
        "age_55_64", "age_65_plus", "housing_type", "location_type",
    ]

    data = {}
    for row in ws.iter_rows(values_only=True):
        if row[0] is None or str(row[0]).strip() in ("ZIP Code", ""):
            continue
        rec = {}
        for i, k in enumerate(keys):
            val = row[i] if i < len(row) else None
            rec[k] = val
        z = str(rec["zip_code"]).zfill(5)
        rec["zip_code"] = z
        data[z] = rec

    wb.close()

    if not data:
        return {"status": "error", "message": "No data found in uploaded file"}

    out_path = SEED_DIR / "census_segments.json"
    with out_path.open("w") as f:
        json.dump(data, f)

    territory = {}
    for z, rec in data.items():
        territory[z] = {
            "city": rec.get("city"),
            "county": rec.get("county"),
            "coverage": rec.get("coverage"),
            "region": rec.get("region"),
        }
    terr_path = DATA_DIR / "territory_zips.json"
    with terr_path.open("w") as f:
        json.dump(territory, f)

    from routers.growth import _build_zip_table, _load_census
    _build_zip_table.cache_clear()
    _load_census.cache_clear()
    _clear_disk_cache()

    return {
        "status": "ok",
        "message": f"Census data updated: {len(data)} ZIPs parsed from '{sheet_name}'",
        "zips": len(data),
    }


# ─── Per-Source Refresh ──────────────────────────────────────────────────

REFRESH_SOURCES = {
    "members": {"fn": "_refresh_members", "label": "Membership (Databricks)", "source": "databricks"},
    "insurance": {"fn": "_refresh_insurance", "label": "Insurance Policies (Databricks)", "source": "databricks"},
    "travel": {"fn": "_refresh_travel", "label": "Travel Transactions 12mo (Databricks)", "source": "databricks"},
    "battery": {"fn": "_refresh_battery", "label": "Battery Tests 12mo (Databricks)", "source": "databricks"},
    "ers": {"fn": "_refresh_ers", "label": "ERS Calls 12mo (Databricks)", "source": "databricks"},
    "ltv": {"fn": "_refresh_ltv", "label": "LTV & Tenure (Databricks)", "source": "databricks"},
    "membership_trend": {"fn": "_refresh_membership_trend", "label": "Membership Trend 5yr (Databricks)", "source": "databricks"},
    "census": {"fn": None, "label": "Census & DMV (Upload)", "source": "upload"},
}


@router.get("/api/growth/refresh-sources")
def list_refresh_sources(_user=Depends(require_admin)):
    """List all available data sources with their last refresh times."""
    sources = []
    for key, info in REFRESH_SOURCES.items():
        if key == "census":
            path = SEED_DIR / "census_segments.json"
        else:
            csv_name = DATASET_CSV_MAP.get(key, "")
            path = DATA_DIR / csv_name if csv_name else Path("/dev/null")

        last_modified = None
        row_count = 0
        if path.exists():
            last_modified = time.strftime("%Y-%m-%d %H:%M", time.localtime(path.stat().st_mtime))
            if path.suffix == ".csv":
                with path.open() as f:
                    row_count = sum(1 for _ in f) - 1
            elif path.suffix == ".json":
                with path.open() as f:
                    row_count = len(json.load(f))

        sources.append({
            "key": key,
            "label": info["label"],
            "source_type": info["source"],
            "last_modified": last_modified,
            "row_count": max(0, row_count),
            "refreshable": info["source"] == "databricks",
        })

    return {"sources": sources}


@router.post("/api/growth/refresh/{source_key}")
def refresh_single_source(source_key: str, background_tasks: BackgroundTasks, _user=Depends(require_admin)):
    """Refresh a single data source by key."""
    if source_key not in REFRESH_SOURCES:
        return {"status": "error", "message": f"Unknown source: {source_key}. Valid: {list(REFRESH_SOURCES.keys())}"}

    info = REFRESH_SOURCES[source_key]
    if info["source"] != "databricks":
        return {"status": "error", "message": f"Source '{source_key}' requires file upload, not API refresh."}

    if _refresh_status["running"]:
        return {"status": "already_running", "message": "A refresh is already in progress"}

    has_databricks = all([
        os.environ.get("DATABRICKS_HOST"),
        os.environ.get("DATABRICKS_TOKEN"),
        os.environ.get("DATABRICKS_HTTP_PATH"),
    ])
    if not has_databricks:
        return {"status": "error", "message": "Databricks credentials not configured"}

    fn_map = {name: fn for name, fn in ALL_DATASETS}

    def _do_single():
        global _refresh_status
        _refresh_status["running"] = True
        _refresh_status["results"] = {}
        _backup_current_data()
        zips = _territory_zips()
        if not zips and source_key not in _NO_ZIP_DATASETS:
            _refresh_status["running"] = False
            _refresh_status["last_error"] = "No territory ZIPs"
            return
        try:
            start_t = time.time()
            fn = fn_map[source_key]
            count = fn() if source_key in _NO_ZIP_DATASETS else fn(zips)
            elapsed = round(time.time() - start_t, 1)
            _refresh_status["results"][source_key] = {"rows": count, "seconds": elapsed, "status": "ok"}
        except Exception as e:
            _refresh_status["results"][source_key] = {"rows": 0, "status": "error", "error": str(e)[:200]}
        _refresh_status["running"] = False
        _refresh_status["last_run"] = date.today().isoformat()
        from routers.growth import _build_zip_table
        _build_zip_table.cache_clear()
        _clear_disk_cache()

    background_tasks.add_task(_do_single)
    return {"status": "started", "message": f"Refreshing '{info['label']}'. Check /api/growth/data-status for progress."}


@router.get("/api/growth/backups")
def list_backups(_user=Depends(require_admin)):
    """List available data backups."""
    backups = []
    if BACKUP_DIR.exists():
        for d in sorted(BACKUP_DIR.iterdir(), reverse=True):
            if d.is_dir():
                files = [f.name for f in d.iterdir()]
                backups.append({
                    "timestamp": d.name,
                    "files": files,
                    "file_count": len(files),
                })
    return {"backups": backups}
