import os
import sys
from dotenv import load_dotenv

sys.path.insert(0, "/Users/abdennourlaaroubi/Library/CloudStorage/OneDrive-EnProIndustriesInc/AAA/Dev/SalesPulse/backend")
load_dotenv("/Users/abdennourlaaroubi/Library/CloudStorage/OneDrive-EnProIndustriesInc/AAA/Dev/SalesPulse/.env", override=True)

import openpyxl
from database import SessionLocal
from models import AdvisorTarget, MonthlyAdvisorTarget
from routers.advisor_targets import _normalize_name
from routers.advisor_targets_excel import _parse_target_value

excel_path = "/Users/abdennourlaaroubi/Downloads/Monthly & Quarterly Branch Performance Dashboard & Incentives.xlsx"
year = 2026

db = SessionLocal()
try:
    print(f"Reading spreadsheet from {excel_path}...")
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb["All Advisors"] if "All Advisors" in wb.sheetnames else wb.active
    print(f"Sheet Name: {ws.title}")
    
    rows = list(ws.iter_rows(values_only=True))
    month_columns = [6, 7, 8, 11, 12, 13, 15, 16, 17, 19, 20, 21]
    month_names = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    
    mismatches = []
    matches = 0
    advisors_checked = 0
    
    R = 1
    total_rows = len(rows)
    while R < total_rows:
        row_data = rows[R]
        raw_name = str(row_data[0] or '').strip()
        if not raw_name:
            R += 6
            continue
            
        sf_name = _normalize_name(raw_name)
        
        # Query advisor target
        adv = db.query(AdvisorTarget).filter(
            (AdvisorTarget.sf_name.ilike(sf_name)) |
            (AdvisorTarget.raw_name.ilike(raw_name))
        ).first()
        
        if not adv:
            print(f"  [MISSING ADVISOR IN DB] Name: '{raw_name}' (Normalized: '{sf_name}')")
            R += 6
            continue
            
        advisors_checked += 1
        
        # Load 12 monthly targets from DB
        db_targets = db.query(MonthlyAdvisorTarget).filter(
            MonthlyAdvisorTarget.advisor_target_id == adv.id,
            MonthlyAdvisorTarget.year == year
        ).order_by(MonthlyAdvisorTarget.month).all()
        
        db_target_map = {r.month: r.target_amount for r in db_targets}
        
        # Compare each month
        advisor_mismatches = []
        for month_idx, col_idx in enumerate(month_columns, 1):
            excel_val = 0.0
            if col_idx < len(row_data):
                excel_val = round(_parse_target_value(row_data[col_idx]))
                
            db_val = round(db_target_map.get(month_idx, 0.0))
            
            if excel_val != db_val:
                advisor_mismatches.append(f"{month_names[month_idx-1]}: Excel={excel_val}, DB={db_val}")
                
        if advisor_mismatches:
            mismatches.append((raw_name, sf_name, advisor_mismatches))
        else:
            matches += 1
            
        R += 6
        
    print(f"\n--- Comparison Finished ---")
    print(f"Advisors checked: {advisors_checked}")
    print(f"Perfect Matches: {matches}")
    print(f"Mismatches: {len(mismatches)}")
    
    if mismatches:
        print("\n--- Detailed Mismatches ---")
        for raw, sf, diffs in mismatches:
            print(f"Advisor: '{raw}' (Salesforce: '{sf}')")
            for d in diffs:
                print(f"  {d}")
                
finally:
    db.close()
