"""Tests for advisor targets Excel import/export endpoints."""

import io
import pytest
from unittest.mock import MagicMock
import openpyxl

from models import AdvisorTarget, MonthlyAdvisorTarget, TargetUpload, User
import cache


def test_export_and_import_targets_excel(api_client, auth_headers, monkeypatch, in_memory_db):
    # Clear both in-memory and disk caches to avoid loading stale data from previous runs
    cache.clear_all(skip_protected=False)

    # Mock Power BI data calls to return consistent results
    mock_travel = MagicMock(return_value=[
        {'name': 'Ashley Gielow', 'branch': 'Amherst', 'date': '2026-01-15', 'commission': 1000.0, 'sales': 10000.0},
        {'name': 'Ashley Gielow', 'branch': 'Amherst', 'date': '2026-02-15', 'commission': 2000.0, 'sales': 20000.0},
    ])
    mock_insurance = MagicMock(return_value=[])
    
    # Mock Salesforce user query for titles
    mock_sf_query = MagicMock(return_value=[
        {'Name': 'Ashley Gielow', 'Title': 'Sr. Travel Advisor'}
    ])

    monkeypatch.setattr('pbi_client.travel_by_advisor_day', mock_travel)
    monkeypatch.setattr('pbi_client.insurance_by_advisor_day', mock_insurance)
    monkeypatch.setattr('sf_client.sf_query_all', mock_sf_query)

    # Pre-seed system auto upload & advisor target
    upload = TargetUpload(
        filename='__sf_auto__',
        line='Travel',
        uploaded_by_id=0,
        uploaded_by_email='system',
        advisor_count=1,
    )
    in_memory_db.add(upload)
    in_memory_db.flush()

    adv_target = AdvisorTarget(
        upload_id=upload.id,
        raw_name='Gielow, Ashley',
        sf_name='Ashley Gielow',
        branch='Amherst',
        title='Sr. Travel Advisor',
        monthly_target=15000.0,
    )
    in_memory_db.add(adv_target)

    # Seed a superadmin user to satisfy require_admin
    from auth import create_token
    import bcrypt
    hashed = bcrypt.hashpw(b'superpass123', bcrypt.gensalt()).decode()
    superadmin = User(
        email='super@nyaaa.com',
        name='Super Admin',
        password_hash=hashed,
        role='superadmin',
        is_active=True
    )
    in_memory_db.add(superadmin)
    in_memory_db.commit()

    token = create_token(superadmin.id, superadmin.email, 'superadmin')
    headers = {'Authorization': f'Bearer {token}'}

    # Diagnostics print
    print("\n--- DB AdvisorTargets in test before request ---")
    for at in in_memory_db.query(AdvisorTarget).all():
        print(f"  ID: {at.id}, sf_name: {at.sf_name}, title: {at.title}, branch: {at.branch}")

    # 1. Verify Excel Export
    resp = api_client.get(
        '/api/targets/monthly/2026/export?line=Travel&base=commission',
        headers=headers
    )
    assert resp.status_code == 200

    print("\n--- DB AdvisorTargets in test after request ---")
    for at in in_memory_db.query(AdvisorTarget).all():
        print(f"  ID: {at.id}, sf_name: {at.sf_name}, title: {at.title}, branch: {at.branch}")
    assert resp.headers['content-type'] == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    
    # Read the generated spreadsheet content
    wb = openpyxl.load_workbook(io.BytesIO(resp.content), data_only=True)
    ws = wb["All Advisors"]
    
    print("Row 2 values:")
    for c in range(1, 10):
        print(f"Col {c}: {ws.cell(row=2, column=c).value}")
        
    # Verify spreadsheet structure
    assert ws.cell(row=1, column=1).value == "Associate"
    assert ws.cell(row=2, column=1).value == "Gielow, Ashley"
    assert ws.cell(row=2, column=2).value == "Sr. Travel Advisor"
    assert ws.cell(row=2, column=3).value == "Amherst"
    
    # Edit the Jan monthly target cell in Excel (Col G, which is Column 7)
    # The first advisor's target row is row 2
    ws.cell(row=2, column=7, value=12500)
    
    # 2. Verify Excel Import
    # Save modified spreadsheet to bytes
    out_buf = io.BytesIO()
    wb.save(out_buf)
    out_buf.seek(0)
    
    files = {
        'file': (
            'Monthly_Targets_Travel_2026.xlsx',
            out_buf,
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    }
    
    import_resp = api_client.post(
        '/api/targets/monthly/2026/import?line=Travel&base=commission',
        files=files,
        headers=headers
    )
    
    assert import_resp.status_code == 200
    data = import_resp.json()
    assert data['status'] == 'success'
    assert data['advisors_updated'] == 1
    assert data['targets_updated'] == 12  # 12 months updated
    
    # Query DB to check if January target was updated to 12500
    jan_target = in_memory_db.query(MonthlyAdvisorTarget).filter(
        MonthlyAdvisorTarget.advisor_target_id == adv_target.id,
        MonthlyAdvisorTarget.year == 2026,
        MonthlyAdvisorTarget.month == 1
    ).first()
    
    assert jan_target is not None
    assert jan_target.target_amount == 12500


def test_save_targets_metadata(api_client, monkeypatch, in_memory_db):
    from auth import create_token
    # Setup test user (superadmin)
    superadmin = in_memory_db.query(User).filter(User.role == 'superadmin').first()
    if not superadmin:
        import bcrypt
        hashed = bcrypt.hashpw(b'superpass123', bcrypt.gensalt()).decode()
        superadmin = User(
            email='super2@nyaaa.com',
            name='Super Admin 2',
            password_hash=hashed,
            role='superadmin',
            is_active=True
        )
        in_memory_db.add(superadmin)
        in_memory_db.commit()

    token = create_token(superadmin.id, superadmin.email, 'superadmin')
    headers = {'Authorization': f'Bearer {token}'}

    # Setup AdvisorTarget
    adv_target = AdvisorTarget(
        upload_id=1,
        raw_name='Smith, Jane',
        sf_name='Jane Smith',
        branch='Buffalo',
        title='Travel Consultant',
        monthly_target=10000.0,
    )
    in_memory_db.add(adv_target)
    in_memory_db.commit()

    # Call PUT /api/admin/targets/monthly with metadata updates
    body = {
        'year': 2026,
        'line': 'Travel',
        'base': 'commission',
        'updates': [
            {
                'advisor_target_id': adv_target.id,
                'months': {'1': 5000, '2': 6000},
                'title': 'Senior Consultant',
                'branch': 'Amherst',
                'monthly_target': 12000.0
            }
        ]
    }
    
    resp = api_client.put('/api/admin/targets/monthly', json=body, headers=headers)
    assert resp.status_code == 200
    
    # Verify updates in DB
    db_adv = in_memory_db.query(AdvisorTarget).filter(AdvisorTarget.id == adv_target.id).first()
    assert db_adv.title == 'Senior Consultant'
    assert db_adv.branch == 'Amherst'
    assert db_adv.monthly_target == 12000.0

