"""
Seed AAA WCNY territory reference data into PostgreSQL (sales schema).

Reads from project root:
  - Battery_Zips.xlsx               → territory_zips table (ZIP, city, county, Full/Partial, region)
  - WCNY_Servicing_County_List.xlsx → territory_counties table (county, Full/Partial)

This is REFERENCE DATA — never cleared by cache operations.
Safe to call repeatedly (idempotent via upsert logic).
"""
import os
import logging
import openpyxl
from sqlalchemy import text
from database import engine, SessionLocal
from models import Base

log = logging.getLogger(__name__)

ROOT             = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..")
BATTERY_ZIPS_PATH = os.path.join(ROOT, "Battery_Zips.xlsx")
COUNTY_LIST_PATH  = os.path.join(ROOT, "WCNY_Servicing_County_List.xlsx")


def seed_territory(force: bool = False):
    """
    Populate territory_zips and territory_counties from xlsx source files.
    Uses upsert logic so duplicate source rows are handled cleanly.

    Args:
        force: If True, truncate tables before re-seeding.
    """
    Base.metadata.create_all(bind=engine, checkfirst=True)

    with engine.connect() as conn:
        zip_count = conn.execute(text("SELECT COUNT(*) FROM territory_zips")).scalar()
        county_count = conn.execute(text("SELECT COUNT(*) FROM territory_counties")).scalar()

        if not force and zip_count > 0 and county_count > 0:
            log.info(f"Territory already seeded ({zip_count} ZIPs, {county_count} counties) — skipping")
            return

        if force:
            conn.execute(text("DELETE FROM territory_zips"))
            conn.execute(text("DELETE FROM territory_counties"))
            conn.commit()
            log.info("Territory tables cleared for re-seed")

        # ── Counties ──────────────────────────────────────────────────────
        if not os.path.exists(COUNTY_LIST_PATH):
            log.warning(f"County list not found: {COUNTY_LIST_PATH}")
        else:
            wb = openpyxl.load_workbook(COUNTY_LIST_PATH, data_only=True)
            rows, seen = [], set()
            for row in wb.active.iter_rows(min_row=2, values_only=True):
                county, service = row[0], row[1]
                if not county or not service:
                    continue
                if str(service).strip().lower() not in ("full", "partial"):
                    continue
                key = str(county).strip().lower()
                if key not in seen:
                    seen.add(key)
                    rows.append({'county': str(county).strip(), 'service_type': str(service).strip()})
            wb.close()

            if rows:
                conn.execute(text(
                    "INSERT INTO territory_counties (county, service_type) "
                    "VALUES (:county, :service_type) "
                    "ON CONFLICT (county) DO UPDATE SET service_type = EXCLUDED.service_type"
                ), rows)
                conn.commit()
                log.info(f"Territory counties seeded: {len(rows)}")

        # ── ZIPs ──────────────────────────────────────────────────────────
        if not os.path.exists(BATTERY_ZIPS_PATH):
            log.warning(f"Battery ZIPs file not found: {BATTERY_ZIPS_PATH}")
        else:
            wb = openpyxl.load_workbook(BATTERY_ZIPS_PATH, data_only=True)
            rows = []
            for row in wb["084 Zip_County_Region"].iter_rows(min_row=2, values_only=True):
                _blank, zip_code, city, county, coverage, region = row
                if not zip_code or not coverage:
                    continue
                if str(coverage).strip().lower() not in ("full", "partial"):
                    continue
                rows.append({
                    'zip_code': str(int(zip_code)).zfill(5),
                    'city': str(city).strip() if city else None,
                    'county': str(county).strip() if county else None,
                    'coverage': str(coverage).strip(),
                    'region': str(region).strip() if region else None,
                })
            wb.close()

            if rows:
                conn.execute(text(
                    "INSERT INTO territory_zips (zip_code, city, county, coverage, region) "
                    "VALUES (:zip_code, :city, :county, :coverage, :region) "
                    "ON CONFLICT (zip_code) DO UPDATE SET "
                    "city = EXCLUDED.city, county = EXCLUDED.county, "
                    "coverage = EXCLUDED.coverage, region = EXCLUDED.region"
                ), rows)
                conn.commit()
                log.info(f"Territory ZIPs seeded: {len(rows)}")

        log.info("Territory seed complete")


if __name__ == "__main__":
    import sys
    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
    seed_territory(force="--force" in sys.argv)
